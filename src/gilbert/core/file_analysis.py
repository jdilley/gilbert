"""Lightweight file metadata analysis for workspace files.

Extracts cheap, header-level stats from common file types so the AI
can see what a file contains without reading the full content. All
analysis is synchronous and designed to be fast — it reads only
headers, first few rows, or stat info. Never loads full file content.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def analyze_file(path: Path, media_type: str) -> dict[str, Any]:
    """Return type-specific metadata for a file.

    Designed to be called from a thread pool executor since it does
    blocking I/O. Returns an empty dict for unrecognized types —
    callers always have size + media_type on the entity anyway.
    """
    try:
        if media_type in ("text/csv", "text/tab-separated-values"):
            return _analyze_csv(path, media_type)
        if media_type.startswith("image/"):
            return _analyze_image(path)
        if media_type == "application/pdf":
            return _analyze_pdf(path)
        if media_type.startswith("text/") or media_type in (
            "application/json",
            "application/xml",
            "application/javascript",
        ):
            return _analyze_text(path)
        if media_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            return _analyze_excel(path)
        return {}
    except Exception:
        logger.debug("file analysis failed for %s", path, exc_info=True)
        return {}


def _analyze_csv(path: Path, media_type: str) -> dict[str, Any]:
    delimiter = "\t" if "tab" in media_type else ","
    result: dict[str, Any] = {}

    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        sniffer_sample = f.read(8192)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sniffer_sample, delimiters=",\t;|")
            delimiter = dialect.delimiter
        except csv.Error:
            pass

        reader = csv.reader(f, delimiter=delimiter)

        # Read header
        try:
            header = next(reader)
        except StopIteration:
            return {"row_count": 0, "column_count": 0}

        result["columns"] = header
        result["column_count"] = len(header)

        # Read sample rows (up to 5)
        sample_rows: list[list[str]] = []
        row_count = 0
        for row in reader:
            row_count += 1
            if len(sample_rows) < 5:
                sample_rows.append(row)

        result["sample_rows"] = sample_rows
        result["row_count"] = row_count

    return result


def _analyze_image(path: Path) -> dict[str, Any]:
    try:
        from PIL import Image

        with Image.open(path) as img:
            return {
                "width": img.width,
                "height": img.height,
                "format": img.format or "",
            }
    except ImportError:
        return {}
    except Exception:
        return {}


def _analyze_pdf(path: Path) -> dict[str, Any]:
    try:
        from pypdf import PdfReader

        reader = PdfReader(path)
        return {"page_count": len(reader.pages)}
    except ImportError:
        pass

    try:
        from PyPDF2 import PdfReader as PdfReader2

        reader2 = PdfReader2(path)
        return {"page_count": len(reader2.pages)}
    except ImportError:
        pass

    return {}


def _analyze_text(path: Path) -> dict[str, Any]:
    line_count = 0
    encoding = "utf-8"
    try:
        with path.open("r", encoding="utf-8", errors="replace") as f:
            for _ in f:
                line_count += 1
    except Exception:
        try:
            with path.open("r", encoding="latin-1") as f:
                for _ in f:
                    line_count += 1
            encoding = "latin-1"
        except Exception:
            return {}

    return {"line_count": line_count, "encoding": encoding}


def _analyze_excel(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        result: dict[str, Any] = {"sheet_names": wb.sheetnames}

        ws = wb.active
        if ws is not None:
            result["row_count"] = ws.max_row or 0
            result["column_count"] = ws.max_column or 0

            # Read column headers from first row
            columns: list[str] = []
            if ws.max_row and ws.max_row > 0:
                for cell in ws[1]:
                    columns.append(str(cell.value) if cell.value is not None else "")
            result["columns"] = columns

        wb.close()
        return result
    except ImportError:
        return {}
    except Exception:
        return {}


def format_metadata_summary(metadata: dict[str, Any], media_type: str) -> str:
    """Format metadata into a short human-readable string for the manifest."""
    parts: list[str] = []

    if "row_count" in metadata and "column_count" in metadata:
        parts.append(
            f"{metadata['row_count']} rows x {metadata['column_count']} cols"
        )
        if metadata.get("columns"):
            cols = metadata["columns"]
            if len(cols) > 6:
                col_str = ", ".join(cols[:6]) + ", ..."
            else:
                col_str = ", ".join(cols)
            parts.append(f"[{col_str}]")

    if "width" in metadata and "height" in metadata:
        parts.append(f"{metadata['width']}x{metadata['height']}")
        if metadata.get("format"):
            parts.append(metadata["format"])

    if "page_count" in metadata:
        pc = metadata["page_count"]
        parts.append(f"{pc} page{'s' if pc != 1 else ''}")

    if "line_count" in metadata and not any(
        k in metadata for k in ("row_count", "page_count")
    ):
        parts.append(f"{metadata['line_count']} lines")

    if "sheet_names" in metadata and len(metadata["sheet_names"]) > 1:
        parts.append(f"{len(metadata['sheet_names'])} sheets")

    return " — ".join(parts) if parts else ""
