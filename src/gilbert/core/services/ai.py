"""AI service — orchestrates AI conversations, tool execution, and persistence.

Also includes internal helpers for persona and user memory (previously
separate services, now merged into AIService).
"""

import asyncio
import contextlib
import contextvars
import json as _json
import logging
import uuid
from datetime import UTC, datetime, timedelta
from typing import Any

from gilbert.core.context import (
    get_current_conversation_id,
    get_current_user,
    set_current_conversation_id,
)
from gilbert.core.services._backend_actions import (
    all_backend_actions,
    invoke_backend_action,
)
from gilbert.core.slash_commands import (
    SlashCommandError,
    extract_command_name,
    format_usage,
    parse_slash_command,
)
from gilbert.interfaces.ai import (
    AIBackend,
    AIContextProfile,
    AIRequest,
    AIResponse,
    ChatTurnResult,
    FileAttachment,
    Message,
    MessageRole,
    ModelInfo,
    StopReason,
    StreamEventType,
)
from gilbert.interfaces.auth import AccessControlProvider, UserContext
from gilbert.interfaces.configuration import (
    ConfigAction,
    ConfigActionResult,
    ConfigParam,
    ConfigurationReader,
)
from gilbert.interfaces.service import Service, ServiceInfo, ServiceResolver
from gilbert.interfaces.storage import (
    Filter,
    FilterOp,
    IndexDefinition,
    Query,
    SortField,
    StorageBackend,
)
from gilbert.interfaces.tools import (
    ToolCall,
    ToolDefinition,
    ToolParameter,
    ToolParameterType,
    ToolProvider,
    ToolResult,
)
from gilbert.interfaces.ui import ToolOutput, UIBlock
from gilbert.interfaces.usage import UsageRecord, UsageRecorder
from gilbert.interfaces.ws import WsConnectionBase

logger = logging.getLogger(__name__)
ai_logger = logging.getLogger("gilbert.ai")

_COLLECTION = "ai_conversations"
_PROFILES_COLLECTION = "ai_profiles"
_ASSIGNMENTS_COLLECTION = "ai_profile_assignments"
_COMPRESSION_STATE_KEY = "compression"
_COMPRESSION_CONFIG_KEY = "compression_config"

_COMPRESSION_SYSTEM_PROMPT = """\
You are a conversation summarizer. Produce a concise, factual summary of the \
conversation history provided.

Preserve:
- Key decisions made and their rationale
- Important facts, data, and specific details (names, dates, numbers)
- User preferences and standing instructions
- Open questions or pending tasks
- Outcomes of tool calls and actions taken
- File attachments: note what was attached (filename, type) and what was \
discussed about it — this is the only record of those files

Omit:
- Pleasantries and small talk
- Redundant back-and-forth
- Raw tool call parameters and verbose tool output (summarize outcomes only)
- Formatting artifacts

If a prior summary is provided under "EXISTING SUMMARY", integrate the new \
messages into it rather than starting from scratch. Output only the summary \
text with no preamble.\
"""

# Attachment limits for chat.message.send frames. Keep these tight enough
# that a single turn can't bloat the conversation entity beyond what
# WebSocket and storage can comfortably round-trip.
_MAX_ATTACHMENTS_PER_MESSAGE = 100
_MAX_IMAGE_BYTES = 5 * 1024 * 1024
_MAX_DOCUMENT_BYTES = 32 * 1024 * 1024  # Anthropic's documented PDF cap.
_MAX_TEXT_BYTES = 512 * 1024  # decoded UTF-8.
# Generic "any file" cap for attachments the AI can't read natively
# (zips, videos, binaries, Office docs that aren't xlsx, …). The bytes
# go to disk via ``POST /api/chat/upload`` and the conversation row
# carries only a workspace reference — so this cap applies at upload
# time, not per-message. 1 GiB ceiling by request.
_MAX_FILE_BYTES = 1024 * 1024 * 1024
# Sum cap across all *inline* bytes in a single chat.message.send
# frame. Reference-mode file attachments don't contribute to this cap
# because they don't ride inside the frame. Kept at 64 MiB so a user
# can still attach a small image + a small text note without friction.
_MAX_TOTAL_ATTACHMENT_BYTES = 64 * 1024 * 1024

# Sentinel appended to the content of an assistant row when the user
# interrupts a turn via ``chat.message.cancel``. Two jobs at once:
#
# 1. AI context: on the NEXT turn, the model sees this exact text in
#    the history and understands the previous response was stopped on
#    purpose and must not be resumed unless the user explicitly asks.
#    Without this sentinel, an empty interrupted assistant row reads as
#    "my unfinished work" and the model happily picks back up.
#
# 2. UI filter: ``_group_persisted_messages_into_turns`` and the live
#    return path both strip any content at or after this marker before
#    setting ``final_content`` on a turn so the directive never leaks
#    into the visible chat bubble. The marker stays in the persisted
#    row (which the AI sees) but not in the rendered one.
_INTERRUPT_MARKER = (
    "[INTERRUPTED BY USER — this response was stopped mid-flight via "
    "the stop button and was NOT completed. Do NOT attempt to continue "
    "or resume this work on the next turn. Treat the user's next "
    "message as a new, standalone request; only revisit this task if "
    "they explicitly ask you to.]"
)


def _empty_round_usage() -> dict[str, Any]:
    """Return a zero-valued per-round usage dict for when no usage is available."""
    return {
        "input_tokens": 0,
        "output_tokens": 0,
        "cache_creation_tokens": 0,
        "cache_read_tokens": 0,
        "cost_usd": 0.0,
    }


def _sum_turn_usage(turn: dict[str, Any]) -> dict[str, Any] | None:
    """Fold per-round and final-round usage into one per-turn total.

    Returns ``None`` when no round in the turn carries usage data — lets the
    chat UI suppress the totals chip entirely for turns that never had
    tokens recorded (e.g. error turns, pre-reporting conversations).
    """
    totals = _empty_round_usage()
    totals["rounds"] = 0
    saw_any = False
    for rnd in turn.get("rounds", []) or []:
        usage = rnd.get("usage")
        if not isinstance(usage, dict):
            continue
        saw_any = True
        totals["input_tokens"] += int(usage.get("input_tokens", 0) or 0)
        totals["output_tokens"] += int(usage.get("output_tokens", 0) or 0)
        totals["cache_creation_tokens"] += int(
            usage.get("cache_creation_tokens", 0) or 0
        )
        totals["cache_read_tokens"] += int(usage.get("cache_read_tokens", 0) or 0)
        totals["cost_usd"] += float(usage.get("cost_usd", 0.0) or 0.0)
        totals["rounds"] += 1
    final = turn.get("final_usage")
    if isinstance(final, dict):
        saw_any = True
        totals["input_tokens"] += int(final.get("input_tokens", 0) or 0)
        totals["output_tokens"] += int(final.get("output_tokens", 0) or 0)
        totals["cache_creation_tokens"] += int(
            final.get("cache_creation_tokens", 0) or 0
        )
        totals["cache_read_tokens"] += int(final.get("cache_read_tokens", 0) or 0)
        totals["cost_usd"] += float(final.get("cost_usd", 0.0) or 0.0)
        totals["rounds"] += 1
    if not saw_any:
        return None
    totals["cost_usd"] = round(totals["cost_usd"], 6)
    return totals


def _strip_interrupt_marker(content: str) -> str:
    """Remove the interrupt sentinel from a content string for display.

    The marker is appended (with a leading blank line) to preserve any
    partial reply text the AI had already streamed before the stop.
    This helper returns just the pre-marker text so the visible chat
    bubble doesn't carry the AI-facing directive.
    """
    if not content:
        return ""
    idx = content.find(_INTERRUPT_MARKER)
    if idx < 0:
        return content
    return content[:idx].rstrip()
_ALLOWED_IMAGE_MEDIA_TYPES = frozenset(
    {
        "image/png",
        "image/jpeg",
        "image/gif",
        "image/webp",
    }
)
_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# Documents we accept on the wire as base64 blobs. PDFs flow through to
# Anthropic's document content block unchanged; xlsx is converted to a
# markdown text attachment at parse time so the AI sees rows it can
# actually reason about.
_ALLOWED_DOCUMENT_MEDIA_TYPES = frozenset(
    {
        "application/pdf",
        _XLSX_MEDIA_TYPE,
    }
)


def _convert_xlsx_to_markdown(data: bytes, name: str) -> str:
    """Render an xlsx workbook as one markdown table per sheet.

    Cells with ``None`` become empty cells; pipes and newlines are
    escaped so the markdown parses. ``load_workbook`` runs in read-only
    mode so memory stays bounded for large sheets.
    """
    import io

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"could not read xlsx {name!r}: {exc}") from exc

    parts: list[str] = [f"# {name}"]
    for sheet in wb.worksheets:
        parts.append("")
        parts.append(f"## Sheet: {sheet.title}")
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            parts.append("_(empty)_")
            continue
        max_cols = max((len(r) for r in rows), default=0)
        if max_cols == 0:
            parts.append("_(empty)_")
            continue

        def _cell(v: Any) -> str:
            if v is None:
                return ""
            return str(v).replace("|", "\\|").replace("\n", " ").replace("\r", " ")

        header_row = rows[0]
        header_cells = [_cell(c) for c in header_row] + [""] * (max_cols - len(header_row))
        parts.append("| " + " | ".join(header_cells) + " |")
        parts.append("| " + " | ".join(["---"] * max_cols) + " |")
        for row in rows[1:]:
            cells = [_cell(c) for c in row] + [""] * (max_cols - len(row))
            parts.append("| " + " | ".join(cells) + " |")

    wb.close()
    return "\n".join(parts) + "\n"


def _parse_frame_attachments(raw: Any) -> list[FileAttachment]:
    """Validate and coerce the ``attachments`` field of a chat.message.send
    frame.

    Each entry must be a dict with ``kind`` in
    ``{"image", "document", "text", "file"}``. The first three are the
    AI-readable kinds (image blocks, PDF/xlsx documents, inlined text);
    ``file`` is a catch-all for anything else the user wants to upload.
    File-kind attachments ride through the conversation row and render
    as download chips on the user's bubble, but the AI only sees a text
    stub naming the file. Raises ``ValueError`` with a user-legible
    message on any validation failure.
    """
    import base64

    if raw is None:
        return []
    if not isinstance(raw, list):
        raise ValueError("attachments must be a list")
    if len(raw) > _MAX_ATTACHMENTS_PER_MESSAGE:
        raise ValueError(
            f"too many attachments (max {_MAX_ATTACHMENTS_PER_MESSAGE})",
        )

    result: list[FileAttachment] = []
    total_bytes = 0

    for idx, item in enumerate(raw):
        if not isinstance(item, dict):
            raise ValueError(f"attachments[{idx}] must be an object")
        kind = str(item.get("kind") or "").lower()
        name = str(item.get("name") or "")
        media_type = str(item.get("media_type") or "").lower()

        if kind == "image":
            data = item.get("data")
            if media_type not in _ALLOWED_IMAGE_MEDIA_TYPES:
                raise ValueError(
                    f"attachments[{idx}] has unsupported image media_type "
                    f"(allowed: {', '.join(sorted(_ALLOWED_IMAGE_MEDIA_TYPES))})",
                )
            if not isinstance(data, str) or not data:
                raise ValueError(
                    f"attachments[{idx}] image data must be a non-empty string",
                )
            try:
                decoded_len = len(base64.b64decode(data, validate=True))
            except Exception as exc:
                raise ValueError(
                    f"attachments[{idx}] has invalid base64: {exc}",
                ) from exc
            if decoded_len > _MAX_IMAGE_BYTES:
                raise ValueError(
                    f"attachments[{idx}] image is too large "
                    f"({decoded_len} bytes > {_MAX_IMAGE_BYTES} max)",
                )
            total_bytes += decoded_len
            result.append(
                FileAttachment(
                    kind="image",
                    name=name,
                    media_type=media_type,
                    data=data,
                )
            )
        elif kind == "document":
            data = item.get("data")
            if not name:
                raise ValueError(f"attachments[{idx}] document requires a name")
            if media_type not in _ALLOWED_DOCUMENT_MEDIA_TYPES:
                raise ValueError(
                    f"attachments[{idx}] has unsupported document media_type "
                    f"(allowed: {', '.join(sorted(_ALLOWED_DOCUMENT_MEDIA_TYPES))})",
                )
            if not isinstance(data, str) or not data:
                raise ValueError(
                    f"attachments[{idx}] document data must be a non-empty string",
                )
            try:
                decoded_bytes = base64.b64decode(data, validate=True)
            except Exception as exc:
                raise ValueError(
                    f"attachments[{idx}] has invalid base64: {exc}",
                ) from exc
            decoded_len = len(decoded_bytes)
            if decoded_len > _MAX_DOCUMENT_BYTES:
                raise ValueError(
                    f"attachments[{idx}] document is too large "
                    f"({decoded_len} bytes > {_MAX_DOCUMENT_BYTES} max)",
                )

            if media_type == _XLSX_MEDIA_TYPE:
                # Excel: decode to a markdown text attachment so the AI
                # sees row data it can reason about. Binary xlsx never
                # hits storage or Anthropic.
                try:
                    markdown = _convert_xlsx_to_markdown(decoded_bytes, name)
                except ValueError:
                    raise
                text_bytes = len(markdown.encode("utf-8"))
                if text_bytes > _MAX_TEXT_BYTES:
                    raise ValueError(
                        f"attachments[{idx}] xlsx converts to too much text "
                        f"({text_bytes} bytes > {_MAX_TEXT_BYTES} max) — "
                        "try a smaller subset",
                    )
                total_bytes += text_bytes
                result.append(
                    FileAttachment(
                        kind="text",
                        name=name,
                        media_type="text/markdown",
                        text=markdown,
                    )
                )
            else:
                total_bytes += decoded_len
                result.append(
                    FileAttachment(
                        kind="document",
                        name=name,
                        media_type=media_type,
                        data=data,
                    )
                )
        elif kind == "text":
            text = item.get("text")
            if not name:
                raise ValueError(f"attachments[{idx}] text requires a name")
            if not isinstance(text, str) or not text:
                raise ValueError(
                    f"attachments[{idx}] text must be a non-empty string",
                )
            byte_len = len(text.encode("utf-8"))
            if byte_len > _MAX_TEXT_BYTES:
                raise ValueError(
                    f"attachments[{idx}] text is too large "
                    f"({byte_len} bytes > {_MAX_TEXT_BYTES} max)",
                )
            total_bytes += byte_len
            result.append(
                FileAttachment(
                    kind="text",
                    name=name,
                    media_type=media_type or "text/plain",
                    text=text,
                )
            )
        elif kind == "file":
            # Arbitrary-file catch-all: the user uploaded something the
            # AI can't natively read (a zip, a video, an executable, a
            # docx, whatever). Two sub-modes:
            #
            # 1. Reference mode (the normal case for anything over a
            #    few MB): the file lives on disk in the chat's
            #    per-conversation workspace, uploaded via
            #    ``POST /api/chat/upload`` — an HTTP endpoint that
            #    streams the bytes directly to disk. The frame
            #    carries only ``workspace_skill`` / ``workspace_path``
            #    / ``workspace_conv`` / ``size`` / ``media_type``, no
            #    base64 data, so the conversation row stays small no
            #    matter how big the file is. The chip renders with a
            #    download link that hits ``GET /api/chat/download``
            #    to stream the file back.
            #
            # 2. Inline mode (legacy / small uploads that bypassed
            #    the upload endpoint): carries base64 ``data``
            #    directly. Capped at ``_MAX_FILE_BYTES`` but subject
            #    to the 1 MiB default WebSocket frame limit in
            #    practice, so this path is only useful for tiny
            #    files. Kept for backward compatibility with any
            #    callers that still pre-base64 their attachments.
            if not name:
                raise ValueError(f"attachments[{idx}] file requires a name")

            workspace_skill = str(item.get("workspace_skill") or "")
            workspace_path = str(item.get("workspace_path") or "")
            workspace_conv = str(item.get("workspace_conv") or "")
            workspace_file_id = str(item.get("workspace_file_id") or "")
            raw_size = item.get("size")
            try:
                reported_size = int(raw_size) if raw_size is not None else 0
            except (TypeError, ValueError):
                raise ValueError(
                    f"attachments[{idx}] size must be an integer",
                ) from None
            if reported_size < 0:
                raise ValueError(
                    f"attachments[{idx}] size must be non-negative",
                )

            if workspace_path:
                # Reference mode: bytes live on disk, not in the
                # frame. Size is the server-reported size from the
                # upload endpoint; we trust it because the upload
                # endpoint is the only thing that fills it in.
                if reported_size > _MAX_FILE_BYTES:
                    raise ValueError(
                        f"attachments[{idx}] file is too large "
                        f"({reported_size} bytes > {_MAX_FILE_BYTES} max)",
                    )
                result.append(
                    FileAttachment(
                        kind="file",
                        name=name,
                        media_type=media_type or "application/octet-stream",
                        workspace_skill=workspace_skill,
                        workspace_path=workspace_path,
                        workspace_conv=workspace_conv,
                        workspace_file_id=workspace_file_id,
                        size=reported_size,
                    )
                )
                continue  # no base64 decode needed, no total_bytes contribution

            # Inline mode fallback.
            data = item.get("data")
            if not isinstance(data, str) or not data:
                raise ValueError(
                    f"attachments[{idx}] file needs either inline data "
                    "or a workspace reference",
                )
            try:
                decoded_len = len(base64.b64decode(data, validate=True))
            except Exception as exc:
                raise ValueError(
                    f"attachments[{idx}] has invalid base64: {exc}",
                ) from exc
            if decoded_len > _MAX_FILE_BYTES:
                raise ValueError(
                    f"attachments[{idx}] file is too large "
                    f"({decoded_len} bytes > {_MAX_FILE_BYTES} max)",
                )
            total_bytes += decoded_len
            result.append(
                FileAttachment(
                    kind="file",
                    name=name,
                    # Browsers leave ``file.type`` empty for many
                    # formats (tar.gz, custom extensions, …). Fall
                    # back to the RFC 2046 default so the persisted
                    # row always carries something meaningful.
                    media_type=media_type or "application/octet-stream",
                    data=data,
                    size=decoded_len,
                )
            )
        else:
            raise ValueError(
                f"attachments[{idx}] has unknown kind {kind!r} "
                "(expected image, document, text, or file)",
            )

        if total_bytes > _MAX_TOTAL_ATTACHMENT_BYTES:
            raise ValueError(
                f"attachments exceed total size cap "
                f"({total_bytes} bytes > {_MAX_TOTAL_ATTACHMENT_BYTES} max)",
            )

    return result


def _serialize_attachments_for_wire(
    attachments: list[FileAttachment],
) -> list[dict[str, Any]]:
    """Shape a list of ``FileAttachment`` for a WebSocket frame payload.

    Inline attachments (``data`` / ``text`` set) round-trip as-is so the
    frontend can render them immediately. Workspace-reference attachments
    carry just their coordinates — the frontend calls
    ``skills.workspace.download`` on click to fetch the bytes, which keeps
    the chat frame small and keeps large generated files (PDFs, images)
    on disk instead of round-tripping through the WS for every history
    load.
    """
    out: list[dict[str, Any]] = []
    for att in attachments:
        entry: dict[str, Any] = {
            "kind": att.kind,
            "name": att.name,
            "media_type": att.media_type,
        }
        if att.data:
            entry["data"] = att.data
        if att.text:
            entry["text"] = att.text
        if att.workspace_skill:
            entry["workspace_skill"] = att.workspace_skill
        if att.workspace_path:
            entry["workspace_path"] = att.workspace_path
        if att.workspace_conv:
            entry["workspace_conv"] = att.workspace_conv
        if att.workspace_file_id:
            entry["workspace_file_id"] = att.workspace_file_id
        if att.size:
            entry["size"] = att.size
        out.append(entry)
    return out


# ── Persona constants and helper ──────────────────────────────

_PERSONA_COLLECTION = "persona"
_PERSONA_ID = "active"

# Default persona shipped with Gilbert
DEFAULT_PERSONA = """\
You are Gilbert, a home and business automation assistant.

## Personality
- Casual, friendly, and professional.
- A bit sarcastic and occasionally funny — but never at the user's expense.
- Keep responses concise. Don't over-explain or narrate what you're doing under the hood.

## Announcements
- When making announcements over speakers after a period of silence, \
open with a brief, natural intro like "Hey team, Gilbert here" or \
"Quick heads up from Gilbert" — vary it each time, keep it fresh, \
don't repeat yourself.
- For rapid follow-up announcements, skip the intro.

## Data & information lookup
- Always check our own project data first before searching the web or \
saying you don't have information. Use project lookup and file search \
tools before falling back to web search.
- When someone asks to see a picture or image of something, first check \
if it matches a project name — then use the project files tool to find \
photos. Only search the web or knowledge base if project files come up empty.
- When someone asks about a person, vehicle, timeline, hours, or status, \
check the synced project data first — it's the most authoritative source \
for anything related to our work.

## Tool use
- When you use a tool, just confirm the result briefly. \
Don't reveal internal details (voice IDs, speaker UIDs, API endpoints, \
credential names, backend types) unless the user specifically asks about configuration.
- If something fails, give a clear, helpful message — not a stack trace.
- Only describe capabilities you actually have tools for. The tools available \
to you depend on the current user's role. If you don't have a tool for \
something, don't mention it at all — not even to say you can't do it. \
Just focus on what you CAN do.\
"""

# Always-appended operational hint. Lives outside ``DEFAULT_PERSONA`` so
# it survives user persona customization — parallel dispatch is a runtime
# affordance, not a personality choice, and losing it silently because
# someone rewrote their persona would be a confusing footgun.
_PARALLEL_TOOL_USE_HINT = """\
## Parallel tool use
When a task needs several tool calls whose inputs don't depend on each \
other's outputs — e.g. announcing on multiple speakers, searching multiple \
topics, fetching multiple URLs, looking up several people — emit all of \
them in a single response. The runtime fans out parallel-safe tools via \
``asyncio.gather`` so they run concurrently; calling them one at a time \
across multiple rounds only makes sense when a later call genuinely needs \
an earlier result.\
"""


class _PersonaHelper:
    """Internal helper — manages the AI persona text in entity storage."""

    def __init__(self, storage: StorageBackend) -> None:
        self._storage = storage
        self._persona: str = DEFAULT_PERSONA
        self._is_customized: bool = False

    async def load(self) -> None:
        saved = await self._storage.get(_PERSONA_COLLECTION, _PERSONA_ID)
        if saved and saved.get("text"):
            self._persona = saved["text"]
            self._is_customized = saved.get("customized", False)
            logger.info("Persona loaded from storage (customized=%s)", self._is_customized)
        else:
            logger.info("No persona stored — using default")

    @property
    def persona(self) -> str:
        return self._persona

    @property
    def is_customized(self) -> bool:
        return self._is_customized

    async def update_persona(self, text: str) -> None:
        self._persona = text
        self._is_customized = True
        await self._storage.put(
            _PERSONA_COLLECTION, _PERSONA_ID, {"text": text, "customized": True}
        )
        logger.info("Persona updated (%d chars)", len(text))

    async def reset_persona(self) -> None:
        self._persona = DEFAULT_PERSONA
        self._is_customized = False
        await self._storage.put(
            _PERSONA_COLLECTION,
            _PERSONA_ID,
            {"text": DEFAULT_PERSONA, "customized": False},
        )
        logger.info("Persona reset to default")


# ── Memory helper ─────────────────────────────────────────────

_MEMORY_COLLECTION = "user_memories"


class _MemoryHelper:
    """Internal helper — per-user persistent memories."""

    def __init__(self, storage: StorageBackend) -> None:
        self._storage = storage

    async def setup_indexes(self) -> None:
        await self._storage.ensure_index(
            IndexDefinition(
                collection=_MEMORY_COLLECTION,
                fields=["user_id"],
            )
        )

    async def get_user_summaries(self, user_id: str) -> str:
        memories = await self._get_user_memories(user_id)
        if not memories:
            return ""
        lines = [f"## Memories for this user ({len(memories)} stored)"]
        for m in memories:
            mid = m.get("_id", "")
            summary = m.get("summary", "")
            source = m.get("source", "user")
            lines.append(f"- [{mid}] {summary} ({source})")
        return "\n".join(lines)

    async def _get_user_memories(self, user_id: str) -> list[dict[str, Any]]:
        memories = await self._storage.query(
            Query(
                collection=_MEMORY_COLLECTION,
                filters=[Filter(field="user_id", op=FilterOp.EQ, value=user_id)],
            )
        )

        def sort_key(m: dict[str, Any]) -> tuple[int, int, str]:
            source_rank = 0 if m.get("source") == "user" else 1
            access = -(m.get("access_count", 0))
            created = m.get("created_at", "")
            return (source_rank, access, created)

        memories.sort(key=sort_key)
        return memories

    async def remember(self, user_id: str, args: dict[str, Any]) -> str:
        summary = args.get("summary", "").strip()
        content = args.get("content", "").strip()
        source = args.get("source", "user")
        if not summary:
            return "I need a summary to remember."
        if not content:
            content = summary
        now = datetime.now(UTC).isoformat()
        memory_id = f"memory_{uuid.uuid4().hex[:12]}"
        await self._storage.put(
            _MEMORY_COLLECTION,
            memory_id,
            {
                "memory_id": memory_id,
                "user_id": user_id,
                "summary": summary,
                "content": content,
                "source": source,
                "access_count": 0,
                "created_at": now,
                "updated_at": now,
            },
        )
        logger.info("Memory created for %s: %s", user_id, summary[:60])
        return f"Got it, I'll remember that. (memory {memory_id})"

    async def recall(self, user_id: str, args: dict[str, Any]) -> str:
        ids: list[str] = args.get("ids", [])
        if not ids:
            return "I need memory IDs to recall. Use 'list' first to see available memories."
        results: list[str] = []
        for mid in ids:
            mid = str(mid)
            record = await self._storage.get(_MEMORY_COLLECTION, mid)
            if not record:
                results.append(f"[{mid}] Not found.")
                continue
            if record.get("user_id") != user_id:
                results.append(f"[{mid}] Not your memory.")
                continue
            record["access_count"] = record.get("access_count", 0) + 1
            await self._storage.put(_MEMORY_COLLECTION, mid, record)
            results.append(
                f"[{mid}] {record.get('summary', '')}\n"
                f"Content: {record.get('content', '')}\n"
                f"Source: {record.get('source', 'user')} | "
                f"Created: {record.get('created_at', '')} | "
                f"Accessed: {record['access_count']} times"
            )
        return "\n\n".join(results)

    async def update(self, user_id: str, args: dict[str, Any]) -> str:
        memory_id = args.get("id", "")
        if not memory_id:
            return "I need a memory ID to update."
        record = await self._storage.get(_MEMORY_COLLECTION, str(memory_id))
        if not record:
            return f"Memory {memory_id} not found."
        if record.get("user_id") != user_id:
            return f"Memory {memory_id} doesn't belong to you."
        summary = args.get("summary")
        content = args.get("content")
        if summary:
            record["summary"] = summary
        if content:
            record["content"] = content
        record["updated_at"] = datetime.now(UTC).isoformat()
        await self._storage.put(_MEMORY_COLLECTION, str(memory_id), record)
        logger.info("Memory updated for %s: %s", user_id, memory_id)
        return f"Memory {memory_id} updated."

    async def forget(self, user_id: str, args: dict[str, Any]) -> str:
        memory_id = args.get("id", "")
        if not memory_id:
            return "I need a memory ID to forget."
        record = await self._storage.get(_MEMORY_COLLECTION, str(memory_id))
        if not record:
            return f"Memory {memory_id} not found."
        if record.get("user_id") != user_id:
            return f"Memory {memory_id} doesn't belong to you."
        await self._storage.delete(_MEMORY_COLLECTION, str(memory_id))
        logger.info("Memory forgotten for %s: %s", user_id, memory_id)
        return f"Memory {memory_id} forgotten."

    async def list_memories(self, user_id: str) -> str:
        memories = await self._get_user_memories(user_id)
        if not memories:
            return "No memories stored for you yet."
        lines = [f"{len(memories)} memory/memories stored:"]
        for m in memories:
            mid = m.get("_id", "")
            summary = m.get("summary", "")
            source = m.get("source", "user")
            access = m.get("access_count", 0)
            lines.append(f"  [{mid}] {summary} ({source}) — accessed {access}x")
        return "\n".join(lines)


# Built-in profiles seeded on first start
_DEFAULT_PROFILE = "standard"

_BUILTIN_PROFILES = [
    AIContextProfile(
        name="light",
        description="Light tier — fast, cost-effective model with all tools",
        tool_mode="all",
    ),
    AIContextProfile(
        name="standard",
        description="Standard tier — balanced model with all tools",
        tool_mode="all",
    ),
    AIContextProfile(
        name="advanced",
        description="Advanced tier — most capable model with all tools",
        tool_mode="all",
    ),
]

_UNDELETABLE_PROFILES = frozenset({"light", "standard", "advanced"})

# Built-in call→profile assignments seeded on first start
_BUILTIN_ASSIGNMENTS: dict[str, str] = {
    "human_chat": "standard",
    "greeting": "light",
    "roast": "standard",
    "scheduled_action": "standard",
    "inbox_ai_chat": "standard",
    "guess_song_validate": "light",
    "mcp_sampling": "standard",
    "mcp_server_client": "standard",
}


class AIService(Service):
    """Orchestrates AI conversations with tool use.

    Wraps an AIBackend (provider-specific) and adds:
    - Agentic loop (tool call → execute → feed back → repeat)
    - Tool discovery from registered ToolProvider services
    - Conversation persistence to storage
    - History truncation
    """

    def __init__(self) -> None:
        self._backends: dict[str, AIBackend] = {}
        self._enabled: bool = False
        self._system_prompt: str = ""
        self._max_history_messages: int = 50
        self._max_tool_rounds: int = 15
        self._max_continuation_rounds: int = 2
        self._compression_enabled: bool = True
        self._compression_threshold: int = 40
        self._compression_keep_recent: int = 20
        self._compression_summary_max_tokens: int = 1500
        self._storage: StorageBackend | None = None
        self._resolver: ServiceResolver | None = None
        self._acl_svc: Any | None = None
        # NOTE: the active conversation id lives in a ContextVar (see
        # gilbert.core.context.{get,set}_current_conversation_id), not
        # on the AIService instance. The service is a singleton shared
        # across users; an instance attribute would race when two
        # conversations overlap (two users, two tabs, shared rooms),
        # causing events to publish with the wrong conv_id and the UI
        # to show tool calls in the wrong chat.
        # AI context profiles
        self._profiles: dict[str, AIContextProfile] = {}
        self._assignments: dict[str, str] = {}  # call_name -> profile_name
        self._default_profile: str = _DEFAULT_PROFILE
        self._chat_profile: str = "standard"
        # Internal helpers (initialized in start())
        self._persona: _PersonaHelper | None = None
        self._memory: _MemoryHelper | None = None
        self._memory_enabled: bool = True
        self._in_flight_chats: dict[str, tuple[asyncio.Task[Any], str]] = {}

    def service_info(self) -> ServiceInfo:
        return ServiceInfo(
            name="ai",
            capabilities=frozenset(
                {"ai_chat", "ai_tools", "ws_handlers", "persona", "user_memory"}
            ),
            requires=frozenset({"entity_storage"}),
            optional=frozenset({"ai_tools", "configuration", "access_control"}),
            events=frozenset({"chat.conversation.renamed"}),
            toggleable=True,
            toggle_description="AI chat and tool execution",
        )

    def get_backend(self, name: str = "") -> AIBackend:
        """Get a backend by name, or the first available one."""
        if name and name in self._backends:
            return self._backends[name]
        if self._backends:
            return next(iter(self._backends.values()))
        raise RuntimeError("No AI backends initialized — service is disabled or not started")

    async def start(self, resolver: ServiceResolver) -> None:
        from gilbert.interfaces.storage import StorageProvider

        # Load tunable config from ConfigurationService if available
        config_svc = resolver.get_capability("configuration")
        section: dict[str, Any] = {}
        if config_svc is not None:
            from gilbert.interfaces.configuration import ConfigurationReader

            if isinstance(config_svc, ConfigurationReader):
                section = config_svc.get_section("ai")
                self._apply_config(section)

        # Check enabled — if False, skip backend init and return early
        if not section.get("enabled", False) and not self._enabled:
            logger.info("AI service disabled")
            return

        self._enabled = True

        # Initialize all configured backends (skip if already injected, e.g. tests)
        if not self._backends:
            await self._reinit_backends(section.get("backends", {}))

        if not self._backends:
            logger.warning("No AI backends initialized — AI will be non-functional")

        # Resolve storage
        storage_svc = resolver.require_capability("entity_storage")
        if not isinstance(storage_svc, StorageProvider):
            raise TypeError("Expected StorageProvider for 'entity_storage' capability")
        self._storage = storage_svc.backend

        # Initialize internal helpers
        self._persona = _PersonaHelper(self._storage)
        await self._persona.load()

        self._memory = _MemoryHelper(self._storage)
        await self._memory.setup_indexes()

        # Resolve access control (optional — if missing, no filtering)
        self._acl_svc = resolver.get_capability("access_control")

        # Save resolver for lazy tool discovery
        self._resolver = resolver

        # Load memory enabled setting
        if config_svc is not None:
            from gilbert.interfaces.configuration import ConfigurationReader

            if isinstance(config_svc, ConfigurationReader):
                memory_section = config_svc.get_section("memory")
                self._memory_enabled = memory_section.get("enabled", True)

        # Load profiles and assignments
        await self._load_profiles()

        logger.info(
            "AI service started (profiles=%d, assignments=%d)",
            len(self._profiles),
            len(self._assignments),
        )

    def _apply_config(self, section: dict[str, Any]) -> None:
        """Apply tunable config values from a config section."""
        self._max_history_messages = section.get("max_history_messages", self._max_history_messages)
        self._max_tool_rounds = section.get("max_tool_rounds", self._max_tool_rounds)
        self._max_continuation_rounds = section.get(
            "max_continuation_rounds", self._max_continuation_rounds
        )
        self._compression_enabled = section.get("compression_enabled", self._compression_enabled)
        self._compression_threshold = section.get(
            "compression_threshold", self._compression_threshold
        )
        self._compression_keep_recent = section.get(
            "compression_keep_recent", self._compression_keep_recent
        )
        self._compression_summary_max_tokens = section.get(
            "compression_summary_max_tokens", self._compression_summary_max_tokens
        )
        self._default_profile = section.get("default_profile", _DEFAULT_PROFILE)
        self._chat_profile = section.get("chat_profile", self._chat_profile)

    # --- Configurable protocol ---

    @property
    def config_namespace(self) -> str:
        return "ai"

    @property
    def config_category(self) -> str:
        return "Intelligence"

    def config_params(self) -> list[ConfigParam]:
        params = [
            ConfigParam(
                key="max_history_messages",
                type=ToolParameterType.INTEGER,
                description="Maximum conversation messages to include in each request.",
                default=50,
            ),
            ConfigParam(
                key="max_tool_rounds",
                type=ToolParameterType.INTEGER,
                description="Maximum agentic loop iterations (tool call rounds) per chat.",
                default=15,
            ),
            ConfigParam(
                key="max_continuation_rounds",
                type=ToolParameterType.INTEGER,
                description=(
                    "When the backend reports a max_tokens cutoff on a text-only "
                    "response, the loop issues a bounded 'please continue' turn. "
                    "This caps how many such recoveries happen per chat turn."
                ),
                default=2,
            ),
            ConfigParam(
                key="default_profile",
                type=ToolParameterType.STRING,
                description=(
                    "Profile used when an AI call has no explicit assignment. "
                    "Typically one of the tier profiles (light, standard, advanced)."
                ),
                default=_DEFAULT_PROFILE,
                choices_from="ai_profiles",
            ),
            ConfigParam(
                key="chat_profile",
                type=ToolParameterType.STRING,
                description="AI profile for web and Slack chat.",
                default="standard",
                choices_from="ai_profiles",
            ),
            ConfigParam(
                key="default_persona",
                type=ToolParameterType.STRING,
                description="Default persona instructions for the AI assistant.",
                default=DEFAULT_PERSONA,
                multiline=True,
            ),
            ConfigParam(
                key="memory_enabled",
                type=ToolParameterType.BOOLEAN,
                description="Whether the AI memory system is enabled.",
                default=True,
                restart_required=True,
            ),
            ConfigParam(
                key="compression_enabled",
                type=ToolParameterType.BOOLEAN,
                description=(
                    "Summarize older messages when a conversation exceeds the "
                    "threshold, preserving context that would otherwise be lost."
                ),
                default=True,
            ),
            ConfigParam(
                key="compression_threshold",
                type=ToolParameterType.INTEGER,
                description=(
                    "Compress when total message count exceeds this value."
                ),
                default=40,
            ),
            ConfigParam(
                key="compression_keep_recent",
                type=ToolParameterType.INTEGER,
                description=(
                    "Always keep this many recent messages verbatim (uncompressed)."
                ),
                default=20,
            ),
            ConfigParam(
                key="compression_summary_max_tokens",
                type=ToolParameterType.INTEGER,
                description=(
                    "Maximum tokens for the AI-generated conversation summary."
                ),
                default=1500,
            ),
        ]
        # Include per-backend config params under backends.<name>.*
        for name, cls in sorted(AIBackend.registered_backends().items()):
            for bp in cls.backend_config_params():
                params.append(
                    ConfigParam(
                        key=f"backends.{name}.{bp.key}",
                        type=bp.type,
                        description=f"[{name}] {bp.description}",
                        default=bp.default,
                        restart_required=bp.restart_required,
                        sensitive=bp.sensitive,
                        choices=bp.choices,
                        choices_from=bp.choices_from,
                        multiline=bp.multiline,
                        backend_param=True,
                    )
                )
        return params

    async def on_config_changed(self, config: dict[str, Any]) -> None:
        self._apply_config(config)
        await self._reinit_backends(config.get("backends", {}))

    async def _reinit_backends(self, backends_config: dict[str, Any]) -> None:
        """Reinitialize backends from config, closing any that changed.

        Backends with ``enabled=False`` in their config section are skipped
        entirely — any existing instance is closed and dropped from the
        registry so profile dropdowns / model lists stop listing them.
        ``enabled`` defaults to True so configs predating the toggle keep
        initializing their backends without manual migration.
        """
        if not isinstance(backends_config, dict):
            return
        for name, cls in AIBackend.registered_backends().items():
            cfg = backends_config.get(name, {})
            if not isinstance(cfg, dict):
                cfg = {}
            enabled = cfg.get("enabled", True) is True
            old = self._backends.get(name)
            if not enabled:
                if old is not None:
                    await old.close()
                    self._backends.pop(name, None)
                    logger.info("AI backend '%s' disabled, closed", name)
                continue
            try:
                inst = cls()
                await inst.initialize(cfg)
                if old is not None:
                    await old.close()
                self._backends[name] = inst
                logger.info("AI backend '%s' (re)initialized", name)
            except Exception as exc:
                if old is not None:
                    pass
                else:
                    logger.debug("AI backend '%s' not ready: %s", name, exc)

    async def stop(self) -> None:
        for backend in self._backends.values():
            await backend.close()
        self._backends.clear()

    # --- ConfigActionProvider ---

    def config_actions(self) -> list[ConfigAction]:
        actions: list[ConfigAction] = []
        for name, backend in self._backends.items():
            actions.extend(
                all_backend_actions(
                    registry={name: type(backend)},
                    current_backend=backend,
                )
            )
        # Also include actions for backends not yet initialized
        for name, cls in AIBackend.registered_backends().items():
            if name not in self._backends:
                actions.extend(
                    all_backend_actions(
                        registry={name: cls},
                        current_backend=None,
                    )
                )
        return actions

    async def invoke_config_action(
        self,
        key: str,
        payload: dict[str, Any],
    ) -> ConfigActionResult:
        backend_name = payload.get("backend", "")
        backend = self._backends.get(backend_name) if backend_name else None

        # If the named backend isn't running, spin up a transient instance
        # from the stored config and run the action on it. Covers two
        # cases the old "fall back to the first live backend" path got
        # wrong: (1) the backend is disabled via its ``enabled`` toggle,
        # and (2) its last init failed (bad api_key, unreachable host)
        # — precisely when the user clicks "Test connection" to diagnose.
        transient: AIBackend | None = None
        if backend is None and backend_name:
            cls = AIBackend.registered_backends().get(backend_name)
            if cls is not None:
                cfg = self._load_backend_config(backend_name)
                try:
                    transient = cls()
                    await transient.initialize(cfg)
                    backend = transient
                except Exception as exc:
                    if transient is not None:
                        with contextlib.suppress(Exception):
                            await transient.close()
                    return ConfigActionResult(
                        status="error",
                        message=(
                            f"Backend '{backend_name}' couldn't initialize "
                            f"with current settings: {exc}"
                        ),
                    )

        if backend is None and self._backends:
            backend = next(iter(self._backends.values()))
        try:
            return await invoke_backend_action(backend, key, payload)
        finally:
            if transient is not None:
                with contextlib.suppress(Exception):
                    await transient.close()

    def _load_backend_config(self, name: str) -> dict[str, Any]:
        """Read ``backends.<name>.*`` from the current AI config section.

        Used to bootstrap a transient backend instance for config actions
        (e.g. ``test_connection``) when the backend isn't currently
        initialized — either because it's disabled, or because its last
        init attempt raised.
        """
        if self._resolver is None:
            return {}
        config_svc = self._resolver.get_capability("configuration")
        if not isinstance(config_svc, ConfigurationReader):
            return {}
        section = config_svc.get_section_safe("ai") or {}
        backends = section.get("backends", {})
        if not isinstance(backends, dict):
            return {}
        cfg = backends.get(name, {})
        return cfg if isinstance(cfg, dict) else {}

    # --- AI Context Profiles ---

    async def _load_profiles(self) -> None:
        """Load profiles and assignments from storage, seeding built-ins on first run.

        Also reconciles stale data: removes legacy built-in profiles that
        are no longer shipped, and updates assignments whose target profile
        was removed or whose built-in default has changed.
        """
        if self._storage is None:
            # No storage — use built-ins in memory only
            self._profiles = {p.name: p for p in _BUILTIN_PROFILES}
            self._assignments = dict(_BUILTIN_ASSIGNMENTS)
            return

        # Seed built-in profiles and assignments only on a fresh database
        # (no profiles exist yet). After that, the user owns the data.
        existing_profiles = await self._storage.query(Query(collection=_PROFILES_COLLECTION))
        if not existing_profiles:
            for bp in _BUILTIN_PROFILES:
                await self._storage.put(
                    _PROFILES_COLLECTION,
                    bp.name,
                    {
                        "name": bp.name,
                        "description": bp.description,
                        "tool_mode": bp.tool_mode,
                        "tools": bp.tools,
                        "tool_roles": bp.tool_roles,
                        "backend": bp.backend,
                        "model": bp.model,
                    },
                )
            for call_name, profile_name in _BUILTIN_ASSIGNMENTS.items():
                await self._storage.put(
                    _ASSIGNMENTS_COLLECTION,
                    call_name,
                    {
                        "call_name": call_name,
                        "profile": profile_name,
                    },
                )

        # Fix any assignment (built-in or user-created) that points to a
        # profile that no longer exists — reset it to the default profile.
        all_assignments = await self._storage.query(Query(collection=_ASSIGNMENTS_COLLECTION))
        for doc in all_assignments:
            target = doc.get("profile", "")
            call_name = doc.get("call_name", "")
            if not target or not call_name:
                continue
            if await self._storage.get(_PROFILES_COLLECTION, target) is None:
                fallback = _BUILTIN_ASSIGNMENTS.get(call_name, self._default_profile)
                await self._storage.put(
                    _ASSIGNMENTS_COLLECTION,
                    call_name,
                    {
                        "call_name": call_name,
                        "profile": fallback,
                    },
                )
                logger.info(
                    "Reset stale assignment '%s' from '%s' to '%s'",
                    call_name, target, fallback,
                )

        # Load all profiles from storage
        await self._refresh_profiles()

    async def _refresh_profiles(self) -> None:
        """Reload profiles and assignments from storage into memory."""
        if self._storage is None:
            return

        # Load profiles
        profile_docs = await self._storage.query(Query(collection=_PROFILES_COLLECTION))
        self._profiles = {}
        for doc in profile_docs:
            name = doc.get("name", "")
            if name:
                self._profiles[name] = AIContextProfile(
                    name=name,
                    description=doc.get("description", ""),
                    tool_mode=doc.get("tool_mode", "all"),
                    tools=doc.get("tools", []),
                    tool_roles=doc.get("tool_roles", {}),
                    backend=doc.get("backend", ""),
                    model=doc.get("model", ""),
                )

        # Load assignments
        assignment_docs = await self._storage.query(Query(collection=_ASSIGNMENTS_COLLECTION))
        self._assignments = {}
        for doc in assignment_docs:
            call_name = doc.get("call_name", "")
            profile = doc.get("profile", "")
            if call_name and profile:
                self._assignments[call_name] = profile

    def get_profile(self, ai_call: str | None) -> AIContextProfile | None:
        """Resolve the profile for an AI call. Returns None if no profile applies."""
        if ai_call is None:
            return None
        profile_name = self._assignments.get(ai_call, self._default_profile)
        return self._profiles.get(profile_name)

    def list_profiles(self) -> list[AIContextProfile]:
        """List all defined profiles."""
        return sorted(self._profiles.values(), key=lambda p: p.name)

    def has_profile(self, name: str) -> bool:
        """Return True if a profile with this name exists."""
        return name in self._profiles

    def list_assignments(self) -> dict[str, str]:
        """List all call→profile assignments."""
        return dict(self._assignments)

    async def set_profile(self, profile: AIContextProfile) -> None:
        """Create or update a profile."""
        if self._storage is not None:
            await self._storage.put(
                _PROFILES_COLLECTION,
                profile.name,
                {
                    "name": profile.name,
                    "description": profile.description,
                    "tool_mode": profile.tool_mode,
                    "tools": profile.tools,
                    "tool_roles": profile.tool_roles,
                    "backend": profile.backend,
                    "model": profile.model,
                },
            )
        self._profiles[profile.name] = profile
        logger.info(
            "Profile '%s' saved (mode=%s, tools=%d)",
            profile.name,
            profile.tool_mode,
            len(profile.tools),
        )

    async def delete_profile(self, name: str) -> None:
        """Delete a profile."""
        if name in _UNDELETABLE_PROFILES:
            raise ValueError(f"Cannot delete the built-in '{name}' profile")
        if name == self._default_profile:
            raise ValueError(f"Cannot delete '{name}' — it is the current default profile")
        if self._storage is not None:
            await self._storage.delete(_PROFILES_COLLECTION, name)
        self._profiles.pop(name, None)
        logger.info("Profile '%s' deleted", name)

    async def set_assignment(self, call_name: str, profile_name: str) -> None:
        """Assign a profile to an AI call."""
        if profile_name not in self._profiles:
            raise ValueError(f"Unknown profile: {profile_name}")
        if self._storage is not None:
            await self._storage.put(
                _ASSIGNMENTS_COLLECTION,
                call_name,
                {
                    "call_name": call_name,
                    "profile": profile_name,
                },
            )
        self._assignments[call_name] = profile_name
        logger.info("Call '%s' assigned to profile '%s'", call_name, profile_name)

    async def clear_assignment(self, call_name: str) -> None:
        """Remove a call→profile assignment (reverts to default)."""
        if self._storage is not None:
            await self._storage.delete(_ASSIGNMENTS_COLLECTION, call_name)
        self._assignments.pop(call_name, None)
        logger.info("Call '%s' assignment cleared", call_name)

    # --- Backend + model resolution ---

    def get_enabled_models(self) -> list[ModelInfo]:
        """Return models from all initialized backends."""
        models: list[ModelInfo] = []
        for backend in self._backends.values():
            models.extend(backend.available_models())
        return models

    def get_backends_with_models(self) -> list[dict[str, Any]]:
        """Return per-backend model lists for the chat UI."""
        result: list[dict[str, Any]] = []
        for name, backend in self._backends.items():
            result.append({
                "name": name,
                "models": [
                    {"id": m.id, "name": m.name, "description": m.description}
                    for m in backend.available_models()
                ],
            })
        return result

    def _resolve_backend_and_model(
        self,
        profile: AIContextProfile | None,
        backend_override: str = "",
        model_override: str = "",
    ) -> tuple[AIBackend, str]:
        """Resolve which backend and model to use for a request.

        Resolution order:
        1. Explicit overrides (from chat UI)
        2. Profile's backend + model fields
        3. First available backend + its default model

        Returns (backend_instance, model_id). model_id may be empty
        (backend uses its own default).
        """
        backend_name = backend_override or (profile.backend if profile else "") or ""
        model = model_override or (profile.model if profile else "") or ""
        backend = self.get_backend(backend_name)
        return backend, model

    # --- One-shot completion (no persistence, no agentic loop) ---

    async def complete_one_shot(
        self,
        *,
        messages: list[Message],
        system_prompt: str = "",
        profile_name: str | None = None,
        max_tokens: int | None = None,
    ) -> AIResponse:
        """Run a single round of the AI backend and return the raw response.

        Unlike ``chat``, this method:

        - Doesn't persist to a conversation.
        - Doesn't loop on tool calls (callers pass a profile with no
          tools if they want that guarantee — this method doesn't
          enforce it).
        - Doesn't take a ``user_ctx`` — ``profile_name`` is the only
          authorization signal, and the caller is expected to have
          already decided the call is safe to make.

        Used by ``MCPService`` to service remote sampling requests.
        Other non-conversational use cases (batch jobs, eval harnesses)
        can adopt the same entry point rather than hand-rolling
        ``AIBackend`` calls.
        """
        if not self._backends:
            raise RuntimeError("No AI backends initialized")
        profile = self._profiles.get(profile_name) if profile_name else None
        backend, model = self._resolve_backend_and_model(profile)
        tools: list[ToolDefinition] = []
        if profile is not None and profile.tool_mode != "include":
            discovered = self._discover_tools(user_ctx=None, profile=profile)
            tools = [td for _, td in discovered.values()]
        request = AIRequest(
            messages=list(messages),
            system_prompt=system_prompt,
            tools=tools,
            model=model,
        )
        response = await backend.generate(request)
        if max_tokens is not None and response.usage is not None:
            # The backend may have respected a different max_tokens;
            # we don't second-guess it, but we surface the usage.
            logger.debug(
                "complete_one_shot used %s tokens (cap was %s)",
                response.usage.input_tokens + response.usage.output_tokens,
                max_tokens,
            )
        # Record usage for one-shot completions (MCP sampling, batch jobs,
        # eval harnesses). Caller is anonymous from the AIService's POV so
        # we stamp it as the SYSTEM user; the profile name + invocation
        # source preserve the context for reporting.
        if response.usage is not None:
            recorder = self._resolve_usage_recorder()
            if recorder is not None:
                try:
                    await recorder.record_round(
                        user_ctx=UserContext.SYSTEM,
                        conversation_id="",
                        profile=profile.name if profile is not None else "",
                        backend=backend.backend_name,
                        model=response.model,
                        usage=response.usage,
                        tool_names=[
                            tc.tool_name for tc in response.message.tool_calls
                        ],
                        stop_reason=response.stop_reason.value,
                        round_num=0,
                        invocation_source="one_shot",
                    )
                except Exception as exc:
                    logger.warning("Usage recorder raised in one_shot: %s", exc)
        return response

    # --- Chat ---

    async def chat(
        self,
        user_message: str,
        conversation_id: str | None = None,
        user_ctx: UserContext | None = None,
        system_prompt: str | None = None,
        ai_call: str | None = None,
        attachments: list[FileAttachment] | None = None,
        model: str = "",
        backend_override: str = "",
        ai_profile: str = "",
    ) -> ChatTurnResult:
        """Send a user message and get an AI response (with full agentic loop).

        Args:
            user_message: The user's input text.
            conversation_id: Existing conversation ID, or None to start new.
            user_ctx: Optional user context. Falls back to contextvar if None.
            system_prompt: Override the system prompt entirely. When ``None``,
                uses the default persona + user memories.
            ai_call: Named AI interaction. Resolved to an AI context profile
                via the assignment table. Ignored when ``ai_profile`` is set.
            attachments: Optional files to attach to this turn's user
                message (images, documents, or text). Backends that
                support multimodal input forward them to the model;
                text attachments are inlined into the prompt body.
            model: Specific model ID override (from chat UI).
            backend: Specific backend override (from chat UI).
            ai_profile: Profile name to use directly, bypassing the
                assignment table.
        """
        if not self._backends:
            raise RuntimeError("No AI backends initialized")
        if user_ctx is None:
            user_ctx = get_current_user()
        # Load or create conversation
        if conversation_id:
            messages = await self._load_conversation(conversation_id)
        else:
            conversation_id = str(uuid.uuid4())
            messages = []

        set_current_conversation_id(conversation_id)

        # ── Slash-command short-circuit ─────────────────────────────
        # If the user typed ``/<name> ...`` and ``<name>`` matches a tool
        # that opted in via ``ToolDefinition.slash_command``, bypass the
        # AI entirely and invoke the tool directly. Grouped commands
        # like ``/radio start`` match a two-word key in the registry;
        # plain ones match a single-word key. Unknown commands are
        # rejected with a helpful error rather than leaked to the AI.
        first_word = extract_command_name(user_message)
        if first_word is not None:
            slash_cmds = self._slash_commands_for_user(user_ctx)
            matched = self._match_slash_command(user_message, slash_cmds)
            if matched is not None:
                return await self._execute_slash_command(
                    user_message,
                    matched,
                    slash_cmds[matched],
                    messages,
                    conversation_id,
                    user_ctx,
                )
            # Unknown slash command — store the attempt and return an
            # actionable error without invoking the AI.
            available = sorted(slash_cmds.keys())
            if available:
                hint = "Available: " + ", ".join(f"/{c}" for c in available)
            else:
                hint = "No slash commands are available to you."
            error_text = f"Unknown slash command '/{first_word}'. {hint}"
            messages.append(
                Message(
                    role=MessageRole.USER,
                    content=user_message,
                    author_id=user_ctx.user_id if user_ctx else "",
                    author_name=user_ctx.display_name if user_ctx else "",
                )
            )
            messages.append(
                Message(
                    role=MessageRole.ASSISTANT,
                    content=error_text,
                )
            )
            await self._save_conversation(
                conversation_id,
                messages,
                user_ctx=user_ctx,
            )
            return ChatTurnResult(
                response_text=error_text,
                conversation_id=conversation_id,
                ui_blocks=[],
                tool_usage=[{"tool_name": f"/{first_word}", "is_error": True}],
                attachments=[],
                rounds=[],
            )

        # Ensure all user attachments are registered in the workspace
        if attachments and user_ctx and self._resolver:
            attachments = await self._ensure_attachments_registered(
                attachments, conversation_id, user_ctx.user_id
            )

        # Append user message
        messages.append(
            Message(
                role=MessageRole.USER,
                content=user_message,
                attachments=list(attachments) if attachments else [],
            )
        )

        # Resolve profile: explicit ai_profile > ai_call assignment > default
        if ai_profile:
            profile = self._profiles.get(ai_profile)
        else:
            profile = self.get_profile(ai_call)

        # Resolve backend + model from profile and overrides
        resolved_backend, resolved_model = self._resolve_backend_and_model(
            profile, backend_override, model
        )
        resolved_backend_name = resolved_backend.backend_name

        # Discover and filter tools based on profile
        tools_by_name = self._discover_tools(user_ctx=user_ctx, profile=profile)

        tool_defs = [defn for _, defn in tools_by_name.values()]

        # Add tools from active skills (additive — only tools that already
        # exist via ToolProviders, restoring any that the profile filtered out)
        if self._resolver:
            skills_svc = self._resolver.get_capability("skills")
            if skills_svc is not None:
                from gilbert.interfaces.skills import SkillsProvider

                if isinstance(skills_svc, SkillsProvider):
                    active = await skills_svc.get_active_skills(conversation_id)
                    if active:
                        skill_tool_names = skills_svc.get_active_allowed_tools(active)
                        if skill_tool_names:
                            # Re-discover unfiltered tools and add missing ones
                            all_tools = self._discover_tools(user_ctx=user_ctx)
                            for tname in skill_tool_names:
                                if tname not in tools_by_name and tname in all_tools:
                                    tools_by_name[tname] = all_tools[tname]
                            tool_defs = [defn for _, defn in tools_by_name.values()]

        # Resolve system prompt — always prepend current date/time
        date_ctx = self._current_datetime_context()
        if system_prompt is not None:
            effective_prompt = f"{date_ctx}\n\n{system_prompt}"
        else:
            effective_prompt = await self._build_system_prompt(
                user_ctx=user_ctx,
                conversation_id=conversation_id,
            )

        # Resolve the list of user_ids that should receive live streaming
        # events for this conversation. Personal chats stream to their
        # owner only; shared rooms stream to every current member. Done
        # once at the start of the turn so we don't hit storage on every
        # text-delta chunk inside the loop.
        stream_visible_to = await self._resolve_stream_audience(
            conversation_id,
            user_ctx,
        )

        # Compress older messages if the conversation is long enough.
        await self._maybe_compress_history(messages, conversation_id)
        compression_state = await self.get_conversation_state(
            _COMPRESSION_STATE_KEY, conversation_id
        )

        # Agentic loop
        response: AIResponse | None = None
        all_ui_blocks: list[UIBlock] = []
        tool_usage: list[dict[str, Any]] = []
        # Set True when the user interrupted this turn mid-flight via
        # ``chat.message.cancel``. The except handler below flips this,
        # marks the trailing assistant row, and then falls through to
        # the shared post-processing (serialize UI blocks, persist, …).
        was_interrupted = False
        # Structured per-round breakdown used by the frontend's turn
        # bubble UI. Each entry represents one AI round that went through
        # tool execution, with the reasoning text the assistant produced
        # alongside the tool_use blocks and the fully-paired tool entries
        # (including final result + is_error). The ``final`` round — the
        # last round of the loop, which produces the user-visible answer
        # — is NOT emitted here; its content rides on ``final_content``
        # on the returned ``ChatTurnResult``.
        turn_rounds: list[dict[str, Any]] = []
        # Aggregate token + cost totals for the whole turn. Each AI round
        # (including the final end_turn round and any max_tokens-recovery
        # continuations) adds into this via ``_record_round_usage``. Ships
        # back to the frontend on ``ChatTurnResult.turn_usage`` so the chat
        # UI can show a per-turn total alongside the answer.
        turn_usage_totals: dict[str, Any] = {
            "input_tokens": 0,
            "output_tokens": 0,
            "cache_creation_tokens": 0,
            "cache_read_tokens": 0,
            "cost_usd": 0.0,
            "rounds": 0,
        }
        # The most-recent round's individual usage + cost. Attached to the
        # next ``turn_rounds`` entry so each round card in the chat UI
        # shows its own token count — not the cumulative running total.
        current_round_usage: dict[str, Any] = _empty_round_usage()
        # Files produced by tool calls during this turn — collected across
        # every round and then landed on the final assistant ``Message`` so
        # the frontend can render download chips next to the reply. Both
        # inline and workspace-reference attachments flow through here.
        turn_attachments: list[FileAttachment] = []
        # Indices (into ``messages``) of synthetic "please continue" user
        # messages injected when a backend reports StopReason.MAX_TOKENS on a
        # text-only response. These exist purely to let the next request
        # continue the reply, and are stripped from the persisted history so
        # the user sees one coherent assistant bubble instead of the internal
        # recovery mechanics.
        continuation_indices: set[int] = set()
        continuation_count = 0

        # Re-assert backend presence and capture a locally-typed
        backend: AIBackend = resolved_backend

        # Label for usage records. ``ai_call`` arg means "invoked from a
        # specific named AI call" (greeting, roast, scheduled action);
        # otherwise it's a normal chat turn.
        invocation_source: str = f"ai_call:{ai_call}" if ai_call else "chat"

        async def _run_agentic_loop() -> None:
            """Execute the agentic for-loop as a cancellable unit.

            Extracted into a nested ``async def`` so the entire loop body
            can be wrapped in a single ``try/except asyncio.CancelledError``
            at the call site without having to re-indent ~250 lines of
            loop logic. The closure captures all outer locals by reference;
            only the two that get rebound (``response`` and
            ``continuation_count``) are declared ``nonlocal``.
            """
            nonlocal response, continuation_count, current_round_usage
            for round_num in range(self._max_tool_rounds):
                truncated = self._truncate_history(
                    messages, compression_state=compression_state
                )

                # Dynamically append conversation state each round so tool-call
                # mutations are visible to subsequent AI rounds.
                conv_state = await self._load_conversation_state(conversation_id)

                # Inject compression summary into the prompt (and exclude
                # the raw compression dict from _format_state_for_context).
                summary_section = ""
                if conv_state:
                    comp = conv_state.pop(_COMPRESSION_STATE_KEY, None)
                    if isinstance(comp, dict) and comp.get("summary"):
                        summary_section = (
                            "\n\n## Prior Conversation Context\n"
                            "The following is a summary of earlier parts of "
                            "this conversation that are no longer shown in "
                            "full:\n\n"
                            + comp["summary"]
                        )
                    conv_state.pop(_COMPRESSION_CONFIG_KEY, None)

                round_prompt = effective_prompt + summary_section
                if conv_state:
                    round_prompt += f"\n\n{self._format_state_for_context(conv_state)}"

                request = AIRequest(
                    messages=truncated,
                    system_prompt=round_prompt,
                    tools=tool_defs if tool_defs else [],
                    model=resolved_model,
                )

                # Drive the backend via ``generate_stream``. For backends that
                # implement true streaming (like AnthropicAI), each TEXT_DELTA
                # chunk gets forwarded onto the event bus as a
                # ``chat.stream.text_delta`` event so the frontend can type
                # out the response live. The MESSAGE_COMPLETE event carries
                # the fully-assembled response that the rest of the agentic
                # loop uses for stop_reason + tool_call handling — identical
                # to the old non-streaming ``await backend.generate()``
                # return value.
                #
                # Backends that don't implement streaming inherit the default
                # fallback on the ABC which calls ``generate()`` and yields
                # exactly one MESSAGE_COMPLETE, so this path is free of cost
                # for them.
                response = None
                async for stream_ev in backend.generate_stream(request):
                    if stream_ev.type == StreamEventType.TEXT_DELTA:
                        if stream_ev.text:
                            await self._publish_event(
                                "chat.stream.text_delta",
                                {
                                    "conversation_id": conversation_id,
                                    "text": stream_ev.text,
                                    "visible_to": stream_visible_to,
                                },
                            )
                    elif stream_ev.type == StreamEventType.MESSAGE_COMPLETE:
                        response = stream_ev.response
                    # TOOL_CALL_START / TOOL_CALL_DELTA / TOOL_CALL_END are
                    # redundant with the chat.tool.started / chat.tool.completed
                    # events that _execute_tool_calls already fires with full
                    # arguments + results. Skip here to avoid double-accounting.

                if response is None:
                    raise RuntimeError(
                        "AI backend stream ended without MESSAGE_COMPLETE — "
                        "this is a backend bug, not a recoverable condition"
                    )
                self._log_api_call(request, response, round_num)

                # Record token usage for this round. The returned dict
                # contains {input_tokens, output_tokens, cache_*_tokens,
                # cost_usd} and is attached to the next ``turn_rounds``
                # entry (for tool rounds) or folded into the turn totals
                # (for the final end_turn round). Never raises — if the
                # UsageService isn't registered yet or storage hiccups,
                # the AI loop continues unaffected.
                current_round_usage = await self._record_round_usage(
                    response=response,
                    user_ctx=user_ctx,
                    conversation_id=conversation_id,
                    profile=profile,
                    backend_name=resolved_backend_name,
                    round_num=round_num,
                    turn_totals=turn_usage_totals,
                    invocation_source=invocation_source,
                )

                # Tell listeners the incremental text for this round is done
                # so they can commit their buffer and prepare for the next
                # round (tool execution, another AI round, or turn end).
                if backend.capabilities().streaming:
                    await self._publish_event(
                        "chat.stream.round_complete",
                        {
                            "conversation_id": conversation_id,
                            "visible_to": stream_visible_to,
                        },
                    )

                # Append assistant message to history. Stamp the round's
                # usage dict onto the message so it persists with the
                # conversation and history replay can surface per-round
                # metrics without a second lookup.
                response.message.usage = dict(current_round_usage)
                messages.append(response.message)

                stop = response.stop_reason

                # Normal tool-use path.
                if stop == StopReason.TOOL_USE and response.message.tool_calls:
                    tool_results, round_ui_blocks = await self._execute_tool_calls(
                        response.message.tool_calls,
                        tools_by_name,
                        user_ctx=user_ctx,
                        profile=profile,
                        backend=backend,
                    )
                    all_ui_blocks.extend(round_ui_blocks)
                    messages.append(Message(role=MessageRole.TOOL_RESULT, tool_results=tool_results))

                    # Collect any files produced by tool calls — these will
                    # ride back on the final assistant ``Message`` at the end
                    # of the turn.
                    for tr in tool_results:
                        if tr.attachments:
                            turn_attachments.extend(tr.attachments)

                    # Track tool usage for the response metadata. Arguments are
                    # sanitized to drop injected ``_user_id`` / ``_room_members``
                    # keys before the payload is sent to the frontend.
                    # The assistant may have emitted a reasoning preamble
                    # alongside its tool_use blocks ("Let me check the
                    # workspace first..."). Attach it to every tool_usage
                    # entry from this round so the frontend's "tools used"
                    # panel can render the reasoning as a caption next to
                    # each call. Persisted with the conversation so a reload
                    # reconstructs the same display.
                    round_reasoning = response.message.content or ""
                    round_tools: list[dict[str, Any]] = []
                    for tc, tr in zip(
                        response.message.tool_calls,
                        tool_results,
                        strict=False,
                    ):
                        entry = {
                            "tool_call_id": tc.tool_call_id,
                            "tool_name": tc.tool_name,
                            "is_error": tr.is_error,
                            "arguments": self._sanitize_tool_args(tc.arguments),
                            "result": tr.content,
                        }
                        # Flat tool_usage retains per-entry reasoning so any
                        # legacy frontend code still works; the new turn UI
                        # reads reasoning off the round instead.
                        tool_usage.append({**entry, "reasoning": round_reasoning})
                        round_tools.append(entry)
                    turn_rounds.append(
                        {
                            "reasoning": round_reasoning,
                            "tools": round_tools,
                            "usage": dict(current_round_usage),
                        }
                    )
                    continue

                # Max-tokens recovery. The backend ran up against its output-token
                # cap before finishing this round. There are two sub-cases:
                #
                # 1. The response carries ``tool_calls`` — we can't tell whether
                #    the tool's JSON input is complete or was cut off mid-field,
                #    and either way executing a partially-specified tool is
                #    unsafe. Strip the tool calls, annotate the message, and
                #    break with an error entry in tool_usage so the frontend can
                #    surface it. Raising ``ai.settings.max_tokens`` is the user-
                #    facing fix, so the annotation tells them that.
                #
                # 2. Text-only — the model ran out of tokens while writing prose.
                #    Inject a synthetic user message asking it to continue, loop
                #    again, and keep doing that up to ``_max_continuation_rounds``
                #    times. The synthetic messages are tracked in
                #    ``continuation_indices`` so they can be stripped before
                #    persistence, and adjacent assistant rows are merged so the
                #    saved history reads as a single coherent reply.
                if stop == StopReason.MAX_TOKENS:
                    if response.message.tool_calls:
                        truncated_names = [tc.tool_name for tc in response.message.tool_calls]
                        logger.warning(
                            "max_tokens truncated a tool call mid-input "
                            "(conversation=%s, tools=%s) — raising "
                            "ai.settings.max_tokens may help",
                            conversation_id,
                            truncated_names,
                        )
                        note = (
                            f"(My previous response was cut off mid tool call "
                            f"({', '.join(truncated_names)}) because it exceeded "
                            f"the model's max_tokens limit. Raise the AI service's "
                            f"max_tokens setting or retry with a smaller request.)"
                        )
                        existing_text = response.message.content or ""
                        combined = (
                            f"{existing_text}\n\n{note}" if existing_text else note
                        )
                        # Rewrite the just-appended assistant row so we don't
                        # persist a broken tool_call input that would make the
                        # next turn's request invalid.
                        messages[-1] = Message(
                            role=MessageRole.ASSISTANT,
                            content=combined,
                            tool_calls=[],
                            author_id=response.message.author_id,
                            author_name=response.message.author_name,
                            visible_to=response.message.visible_to,
                            attachments=list(response.message.attachments),
                            usage=response.message.usage,
                        )
                        response = AIResponse(
                            message=messages[-1],
                            model=response.model,
                            stop_reason=StopReason.MAX_TOKENS,
                            usage=response.usage,
                        )
                        tool_usage.append(
                            {
                                "tool_name": "<max_tokens_truncation>",
                                "is_error": True,
                                "arguments": {},
                                "result": f"Truncated mid tool_use: {', '.join(truncated_names)}",
                            }
                        )
                        break

                    # Text-only truncation — bounded continuation.
                    if continuation_count >= self._max_continuation_rounds:
                        logger.warning(
                            "max_tokens recovery exhausted after %d continuations "
                            "(conversation=%s)",
                            continuation_count,
                            conversation_id,
                        )
                        existing_text = response.message.content or ""
                        annotated = (
                            existing_text
                            + "\n\n(Note: response still truncated after "
                            f"{continuation_count} continuation attempts. Raise "
                            "the AI service's max_tokens and retry for a "
                            "complete reply.)"
                        )
                        messages[-1] = Message(
                            role=MessageRole.ASSISTANT,
                            content=annotated,
                            author_id=response.message.author_id,
                            author_name=response.message.author_name,
                            visible_to=response.message.visible_to,
                            attachments=list(response.message.attachments),
                            usage=response.message.usage,
                        )
                        response = AIResponse(
                            message=messages[-1],
                            model=response.model,
                            stop_reason=StopReason.MAX_TOKENS,
                            usage=response.usage,
                        )
                        break

                    continuation_count += 1
                    continuation_indices.add(len(messages))
                    messages.append(
                        Message(
                            role=MessageRole.USER,
                            content=(
                                "Please continue your previous response — it was "
                                "cut off by a response size limit. Pick up "
                                "exactly where you left off; do not repeat "
                                "anything you've already said."
                            ),
                        )
                    )
                    continue

                # END_TURN or any other terminal stop — normal completion.
                break
            else:
                logger.warning(
                    "Agentic loop hit max rounds (%d) for conversation %s",
                    self._max_tool_rounds,
                    conversation_id,
                )

        try:
            await _run_agentic_loop()
        except asyncio.CancelledError:
            # User hit the stop button (``chat.message.cancel`` RPC).
            # Cancellation propagates through the running stream / tool
            # await, unwinding here with whatever state the loop had
            # accumulated so far: completed rounds in ``turn_rounds``,
            # completed tool results in ``messages``, partial reasoning
            # on the last assistant row (if any). Mark the trailing
            # assistant row as interrupted so the frontend can render
            # the subtle stop indicator, then fall through to the
            # shared post-processing which persists what we've got.
            #
            # IMPORTANT: we also set a clear directive on the content
            # field of the interrupted assistant row. Without this, on
            # the next turn the AI sees an empty (or partial) assistant
            # row and interprets it as "my unfinished work — let me
            # resume that" even when the user's new message is a fresh
            # ask. The explicit "do not resume" sentence stops that
            # behavior cold and is visible to the AI on every
            # subsequent turn's history replay.
            was_interrupted = True
            logger.info(
                "Chat turn interrupted by user for conversation %s "
                "after %d round(s)",
                conversation_id,
                len(turn_rounds),
            )
            if messages and messages[-1].role == MessageRole.ASSISTANT:
                existing = messages[-1].content or ""
                messages[-1].interrupted = True
                messages[-1].content = (
                    f"{existing}\n\n{_INTERRUPT_MARKER}"
                    if existing
                    else _INTERRUPT_MARKER
                )
            else:
                # No trailing assistant row (e.g. cancelled during the
                # very first stream read before MESSAGE_COMPLETE). Add
                # a placeholder so the turn has a detectable stopping
                # point and ``finalize_current_turn`` in turn grouping
                # doesn't mis-flag it. The content carries the explicit
                # directive for the AI.
                messages.append(
                    Message(
                        role=MessageRole.ASSISTANT,
                        content=_INTERRUPT_MARKER,
                        interrupted=True,
                    )
                )

        # Collapse any max_tokens continuations: drop the synthetic
        # continuation user rows and merge adjacent assistant text rows so
        # both the persisted history and the returned final_text look like
        # a single coherent reply instead of leaking the recovery mechanics.
        if continuation_indices:
            messages[:] = self._collapse_continuations(
                messages,
                continuation_indices,
            )
            # The final assistant row after collapse carries the concatenated
            # text. Rebuild ``response`` so the tuple returned to the caller
            # reflects that combined content.
            if messages and messages[-1].role == MessageRole.ASSISTANT and response is not None:
                response = AIResponse(
                    message=messages[-1],
                    model=response.model,
                    stop_reason=response.stop_reason,
                    usage=response.usage,
                )

        # Land any tool-produced attachments on the final assistant
        # ``Message`` so they get persisted with the conversation and ride
        # through the WS send result. We mutate the last assistant row in
        # place — ``Message`` is a mutable dataclass, and its
        # ``attachments`` field is a list.
        if turn_attachments:
            for msg in reversed(messages):
                if msg.role == MessageRole.ASSISTANT:
                    msg.attachments.extend(turn_attachments)
                    # Rebuild ``response`` so the live return carries them too.
                    if response is not None and response.message is not msg:
                        response = AIResponse(
                            message=msg,
                            model=response.model,
                            stop_reason=response.stop_reason,
                            usage=response.usage,
                        )
                    break

        # ``response_index`` anchors UI blocks to the turn they were
        # produced in. The frontend renders one ``TurnBubble`` per
        # user→assistant exchange (one turn per user message), so we
        # count user rows in the persisted history and use that as the
        # turn index. This matches ``MessageList``'s
        # ``assistantToTurnIndex`` map exactly — both sides agree on
        # "the Nth turn is the Nth user message."
        #
        # Counting visible-content assistant rows (the old approach)
        # broke for any turn that had multiple intermediate rounds
        # with reasoning text — every round counted as its own
        # "visible assistant" on the backend, but the frontend's
        # turn-grouped view collapsed all those rounds into one turn.
        # The indices drifted apart and UI blocks fell through to the
        # unanchored bucket at the bottom of the chat.
        user_count = sum(1 for m in messages if m.role == MessageRole.USER)
        response_index = max(0, user_count - 1)

        # Serialize UI blocks with position and submission state
        ui_block_dicts: list[dict[str, Any]] = []
        for block in all_ui_blocks:
            d = block.to_dict()
            d["response_index"] = response_index
            d["submitted"] = False
            d["submission"] = None
            ui_block_dicts.append(d)

        # Persist conversation with user ownership and UI blocks.
        # When the turn was interrupted, shield the save so a
        # subsequent ``task.cancel()`` from the caller (or any stray
        # cancellation already queued on the event loop) can't abort
        # persistence and lose the partial state we just captured.
        save_coro = self._save_conversation(
            conversation_id,
            messages,
            user_ctx,
            ui_blocks=ui_block_dicts,
        )
        if was_interrupted:
            await asyncio.shield(save_coro)
        else:
            await save_coro

        # Return final text response. When the turn was interrupted
        # the trailing assistant row's content carries the AI-facing
        # ``_INTERRUPT_MARKER`` sentinel; strip it before handing the
        # text back to the frontend so the visible bubble shows only
        # whatever partial reply the user actually saw stream in
        # (which may be empty).
        if was_interrupted and messages and messages[-1].role == MessageRole.ASSISTANT:
            final_text = _strip_interrupt_marker(messages[-1].content or "")
        else:
            final_text = response.message.content if response else ""

        # Signal end-of-turn to streaming listeners so they can drop
        # any in-flight streaming buffers and fall back to the
        # authoritative send.result/history paths. No-op when the
        # backend doesn't support streaming.
        if backend.capabilities().streaming:
            await self._publish_event(
                "chat.stream.turn_complete",
                {
                    "conversation_id": conversation_id,
                    "visible_to": stream_visible_to,
                },
            )

        return ChatTurnResult(
            response_text=final_text,
            conversation_id=conversation_id,
            ui_blocks=ui_block_dicts,
            tool_usage=tool_usage,
            attachments=list(turn_attachments),
            rounds=turn_rounds,
            interrupted=was_interrupted,
            model=response.model if response else resolved_model,
            turn_usage=dict(turn_usage_totals),
        )

    # --- System Prompt ---

    @staticmethod
    def _current_datetime_context() -> str:
        """Build a date/time context string in Los Angeles timezone."""
        try:
            from zoneinfo import ZoneInfo

            now = datetime.now(ZoneInfo("America/Los_Angeles"))
        except Exception:
            now = datetime.now(UTC)
        today = now.strftime("%A, %B %d, %Y")
        time_str = now.strftime("%I:%M %p %Z")
        yesterday = (now - timedelta(days=1)).strftime("%A, %B %d, %Y")
        return f"Current date and time: {today} at {time_str}. Yesterday was {yesterday}."

    async def _build_system_prompt(
        self,
        user_ctx: UserContext | None = None,
        conversation_id: str | None = None,
    ) -> str:
        """Build the full system prompt: base identity, persona, user memories, and active skills."""
        parts: list[str] = []

        # Always inject current date/time first
        parts.append(self._current_datetime_context())

        if self._system_prompt:
            parts.append(self._system_prompt)
        if self._persona is not None:
            parts.append(self._persona.persona)
            if not self._persona.is_customized:
                parts.append(
                    "IMPORTANT: The persona has not been customized yet. "
                    "At the start of the FIRST conversation only, briefly let the user know "
                    "they can customize your personality and behavior by asking you to "
                    "update the persona. Only mention this once — never bring it up again "
                    "in subsequent messages or conversations."
                )

        # Always append the parallel-tool-use hint so the model knows it
        # should batch independent tool calls into one response. Lives
        # here rather than inside the persona so a customized persona
        # doesn't accidentally drop it.
        parts.append(_PARALLEL_TOOL_USE_HINT)

        # Inject the active user's identity so the AI knows who it's
        # talking to without having to ask. Skipped for system/guest
        # callers (no real identity) and shared rooms (multiple users
        # in flight; the per-message [Name]: prefix already attributes
        # each utterance). When this is set, the AI can reference the
        # user by name, address email "to the user" using the email
        # below, and avoid asking questions like "what's your email?"
        # that the user record already answers.
        if user_ctx and user_ctx.user_id not in ("system", "guest"):
            identity_lines: list[str] = ["## You're talking to"]
            if user_ctx.display_name:
                identity_lines.append(f"- **Name:** {user_ctx.display_name}")
            if user_ctx.email:
                identity_lines.append(f"- **Email:** {user_ctx.email}")
            identity_lines.append(f"- **User ID:** {user_ctx.user_id}")
            identity_lines.append(
                "Use these details when the request implies the user themself "
                "as the recipient, sender, or subject (e.g. 'email me the file' "
                "→ send to the email above, 'what's my schedule' → look up "
                "this user's id). Don't ask the user for information already "
                "listed here."
            )
            parts.append("\n".join(identity_lines))

        # Inject user memory summaries if available
        if user_ctx and user_ctx.user_id not in ("system", "guest"):
            if self._memory is not None and self._memory_enabled:
                try:
                    summaries = await self._memory.get_user_summaries(user_ctx.user_id)
                    if summaries:
                        parts.append(summaries)
                except Exception:
                    pass  # Memory unavailable — not critical

        # Inject skill system awareness and active skill instructions
        if self._resolver:
            skills_svc = self._resolver.get_capability("skills")
            if skills_svc is not None:
                parts.append(
                    "## Skills\n"
                    "This system supports skills — specialized instruction sets that "
                    "users can enable or disable per conversation. Skills may appear or "
                    "disappear between messages as the user toggles them. When skills "
                    "are active for this conversation, their instructions will appear "
                    "below. Follow them when relevant. If a skill you were using "
                    "disappears, the user disabled it — stop following its instructions.\n\n"
                    "**Important — gated skill access.** Skill tools (read_skill_file, "
                    "run_skill_script) only work on skills the user has activated for "
                    "THIS conversation. If you reach for a skill that isn't active, "
                    "the tool will refuse with an instruction telling you to ask the "
                    "user to enable it. Do NOT call those tools speculatively to see "
                    "what skills exist — only use skills that already appear below. "
                    "If you need a skill that isn't active, ask the user to enable it "
                    "from the Skills panel and then ask you again.\n\n"
                    "**Workspace tools** (browse_workspace, read_workspace_file, "
                    "write_workspace_file, run_workspace_script, attach_workspace_file) "
                    "operate on the conversation's shared file workspace — uploads from "
                    "the user land in uploads/, scripts and temp files go in scratch/, "
                    "and deliverables go in outputs/. These do NOT require skill "
                    "activation.\n\n"
                    "### Creating Skills\n"
                    "Users can ask you to create custom skills. When they do, guide them "
                    "through the process conversationally — you don't need to explain the "
                    "SKILL.md format to them. Instead:\n"
                    "1. Ask what the skill should help with — its purpose and when it should be used.\n"
                    "2. Ask about the specific steps, workflows, or guidelines it should follow.\n"
                    "3. Ask about any gotchas, edge cases, or important constraints.\n"
                    "4. Once you have enough information, use the `create_skill` tool to create it.\n\n"
                    "Scope: By default, create skills as personal (scope='user'). Only offer "
                    "to create a global skill if the user explicitly asks for it — the system "
                    "will enforce permissions automatically. Do NOT ask about scope unless "
                    "the user brings it up.\n\n"
                    "When building the SKILL.md content for `create_skill`:\n"
                    "- The frontmatter MUST include `name` (kebab-case, e.g. 'sales-outreach') "
                    "and `description` (1-2 sentences explaining what it does and when to use it).\n"
                    "- Optionally include `metadata.category` and `metadata.icon` for UI grouping.\n"
                    "- Optionally include `allowed-tools` (space-separated tool names) to declare "
                    "which tools the skill uses — these are existing tools, NOT scripts.\n"
                    "- The body should contain clear, actionable instructions: workflows, "
                    "decision trees, gotchas, templates, and examples.\n"
                    "- Entity-stored skills CANNOT execute scripts or read files from disk. "
                    "They CAN use any AI tools available in the conversation (search, "
                    "data lookups, web fetch, etc.).\n"
                    "- Keep the instructions focused and under 500 lines.\n"
                    "- After creating, let the user know they can enable it from the Skills "
                    "panel in chat settings."
                )
                if conversation_id:
                    try:
                        from gilbert.interfaces.skills import SkillsProvider

                        if isinstance(skills_svc, SkillsProvider):
                            skills_ctx = await skills_svc.build_skills_context(
                                conversation_id,
                            )
                            if skills_ctx:
                                parts.append(skills_ctx)
                    except Exception:
                        pass  # Skills unavailable — not critical

        # Inject workspace file manifest so the AI knows what files are
        # available and their metadata (row counts, dimensions, etc.)
        if self._resolver and conversation_id:
            try:
                from gilbert.interfaces.workspace import WorkspaceProvider

                ws_svc = self._resolver.get_capability("workspace")
                if isinstance(ws_svc, WorkspaceProvider):
                    manifest = await ws_svc.build_workspace_manifest(
                        conversation_id,
                    )
                    if manifest:
                        parts.append(manifest)
                        parts.append(
                            "## File Handling Guidelines\n"
                            "- **Reading small files** (read_workspace_file): text under "
                            "~50 KB or ~100 rows, small JSON/XML\n"
                            "- **Everything else** (run_workspace_script): large files, "
                            "binary formats, aggregations, filtering, joins, "
                            "image processing, XLSX, CSV analysis\n"
                            "- **Generating output files**: ALWAYS use run_workspace_script "
                            "to write data files (CSVs, reports, charts). The script writes "
                            "to disk directly, avoiding token limits. NEVER put large data "
                            "in write_workspace_file content — it has a 512 KB cap and "
                            "wastes tokens. write_workspace_file is only for small scripts "
                            "and config files.\n"
                            "- **File paths**: use flat filenames (e.g. 'analysis.csv', "
                            "'compare.py'), not nested subdirectories\n"
                            "- After generating output, call attach_workspace_file to "
                            "make it downloadable, then annotate_workspace_file for "
                            "description and lineage\n"
                            "- To delete files, use delete_workspace_file"
                        )
            except Exception:
                pass

        return "\n\n".join(parts) if parts else ""


    async def _ensure_attachments_registered(
        self,
        attachments: list[FileAttachment],
        conversation_id: str,
        user_id: str,
    ) -> list[FileAttachment]:
        """Ensure all user attachments are registered in the workspace.

        All attachments now arrive as workspace references (uploaded via
        POST /api/chat/upload). This method checks that each one has a
        registry entry and creates one if missing (defensive — the upload
        endpoint should have already registered it).
        """
        from gilbert.interfaces.workspace import WorkspaceProvider

        ws_svc = self._resolver.get_capability("workspace") if self._resolver else None
        if not isinstance(ws_svc, WorkspaceProvider):
            return attachments

        for att in attachments:
            if att.is_reference and not att.workspace_file_id and att.workspace_path:
                try:
                    existing = await ws_svc.list_files(conversation_id, "upload")
                    found = any(
                        f.get("rel_path") == att.workspace_path for f in existing
                    )
                    if not found:
                        await ws_svc.register_file(
                            conversation_id=conversation_id,
                            user_id=user_id,
                            category="upload",
                            filename=att.name or att.workspace_path.split("/")[-1],
                            rel_path=att.workspace_path,
                            media_type=att.media_type,
                            size=att.size,
                            created_by="user",
                        )
                except Exception:
                    pass

        return attachments

    # --- Tool Discovery ---

    def discover_tools(
        self,
        *,
        user_ctx: UserContext,
        profile_name: str | None = None,
    ) -> dict[str, tuple[ToolProvider, ToolDefinition]]:
        """Public entry point for non-chat callers that want a filtered
        tool list (profile + RBAC applied).

        Used by the MCP server endpoint in Part 4.2 — it builds the
        tool set exposed to external MCP clients. Takes a profile
        *name* rather than a profile object so the caller doesn't
        need to resolve profiles itself. An unknown profile name is
        treated as "no profile" (same as ``profile_name=None``),
        matching how the ``ai_call`` parameter on ``chat`` handles
        unassigned call names.
        """
        profile: AIContextProfile | None = None
        if profile_name:
            profile = self._profiles.get(profile_name)
            if profile is None:
                logger.warning(
                    "discover_tools: unknown profile %r, falling back to all tools",
                    profile_name,
                )
        return self._discover_tools(user_ctx=user_ctx, profile=profile)

    def _discover_tools(
        self,
        user_ctx: UserContext | None = None,
        profile: AIContextProfile | None = None,
    ) -> dict[str, tuple[ToolProvider, ToolDefinition]]:
        """Find all started services that implement ToolProvider and collect their tools.

        If a *profile* is provided, tools are filtered by its tool_mode:
        - ``all``: all tools (RBAC still applies)
        - ``include``: only tools named in ``profile.tools``
        - ``exclude``: all tools except those named in ``profile.tools``

        If the profile defines ``tool_roles``, those override each tool's
        ``required_role`` for RBAC checks within this call.
        """
        tools_by_name: dict[str, tuple[ToolProvider, ToolDefinition]] = {}
        if self._resolver is None:
            return tools_by_name

        for svc in self._resolver.get_all("ai_tools"):
            if not isinstance(svc, ToolProvider):
                continue
            for tool_def in svc.get_tools(user_ctx):
                if tool_def.name in tools_by_name:
                    logger.warning(
                        "Duplicate tool name %r from %s (already registered by %s)",
                        tool_def.name,
                        svc.tool_provider_name,
                        tools_by_name[tool_def.name][0].tool_provider_name,
                    )
                    continue
                tools_by_name[tool_def.name] = (svc, tool_def)

        # Apply profile tool filtering
        if profile is not None:
            if profile.tool_mode == "include":
                include_set = set(profile.tools)
                tools_by_name = {
                    name: v for name, v in tools_by_name.items() if name in include_set
                }
            elif profile.tool_mode == "exclude":
                exclude_set = set(profile.tools)
                tools_by_name = {
                    name: v for name, v in tools_by_name.items() if name not in exclude_set
                }
            # "all" = no filtering

        # Apply RBAC permissions (with optional profile role overrides)
        if user_ctx is not None and self._acl_svc is not None:
            if isinstance(self._acl_svc, AccessControlProvider):
                tool_roles = profile.tool_roles if profile else {}
                before = len(tools_by_name)
                filtered: dict[str, tuple[ToolProvider, ToolDefinition]] = {}
                for name, (prov, tdef) in tools_by_name.items():
                    # Use profile role override if present, else tool's default
                    effective_role = tool_roles.get(name, tdef.required_role)
                    role_level = self._acl_svc.get_role_level(effective_role)
                    user_level = self._acl_svc.get_effective_level(user_ctx)
                    if user_level <= role_level:
                        filtered[name] = (prov, tdef)
                removed = before - len(filtered)
                if removed:
                    logger.debug(
                        "Filtered %d tools for user %s (effective level %d)",
                        removed,
                        user_ctx.user_id,
                        self._acl_svc.get_effective_level(user_ctx),
                    )
                tools_by_name = filtered

        return tools_by_name

    # --- Tool Execution ---

    async def _execute_tool_calls(
        self,
        tool_calls: list[ToolCall],
        tools_by_name: dict[str, tuple[ToolProvider, ToolDefinition]],
        user_ctx: UserContext | None = None,
        profile: AIContextProfile | None = None,
        backend: AIBackend | None = None,
    ) -> tuple[list[ToolResult], list[UIBlock]]:
        """Execute a batch of tool calls and return results + any UI blocks.

        Tools marked ``parallel_safe=True`` run concurrently via
        ``asyncio.gather`` when the active backend advertises
        ``parallel_tool_calls=True``. Unsafe tools — or any tool on a
        backend without the capability — run one at a time, preserving
        the original serial behavior. Result order always matches the
        input ``tool_calls`` order regardless of completion order.
        """
        tool_roles = profile.tool_roles if profile else {}
        backend_parallel = bool(
            backend is not None and backend.capabilities().parallel_tool_calls
        )

        # Pre-resolve room membership once per batch — it's stable across
        # the turn, so we don't need to re-fetch it per tool call (which
        # would also serialize storage reads).
        conv_id = get_current_conversation_id()
        room_members: list[dict[str, str]] | None = None
        storage = getattr(self, "_storage", None)
        if conv_id and storage:
            conv_data = await storage.get(_COLLECTION, conv_id)
            if conv_data and conv_data.get("shared"):
                room_members = [
                    {
                        "user_id": m.get("user_id", ""),
                        "display_name": m.get("display_name", ""),
                    }
                    for m in conv_data.get("members", [])
                ]

        # Partition into execution groups. A "group" is either a single
        # tool (serial path) or a run of consecutive parallel_safe tools
        # (gather path). Consecutive grouping preserves the model's
        # intended ordering: if the model emits [read, write, read], the
        # two reads don't fan across the write.
        groups: list[list[ToolCall]] = []
        for tc in tool_calls:
            pd = tools_by_name.get(tc.tool_name)
            is_parallel = (
                backend_parallel and pd is not None and pd[1].parallel_safe
            )
            if is_parallel and groups and self._group_is_parallel(
                groups[-1], tools_by_name
            ):
                groups[-1].append(tc)
            else:
                groups.append([tc])

        results: list[ToolResult] = []
        ui_blocks: list[UIBlock] = []

        for group in groups:
            if len(group) == 1:
                tr, blocks = await self._run_one_tool(
                    group[0],
                    tools_by_name,
                    user_ctx=user_ctx,
                    tool_roles=tool_roles,
                    conv_id=conv_id,
                    room_members=room_members,
                )
                results.append(tr)
                ui_blocks.extend(blocks)
                continue

            # Parallel batch: each task gets its own contextvars copy so
            # ``set_current_user`` / any other ContextVar set inside one
            # task cannot bleed into its siblings. ``asyncio.Task`` with
            # an explicit ``context=`` is the supported way to do that
            # (Python 3.11+; we're on 3.12).
            logger.debug(
                "Running %d tools in parallel: %s",
                len(group),
                ", ".join(tc.tool_name for tc in group),
            )
            tasks = [
                asyncio.Task(
                    self._run_one_tool(
                        tc,
                        tools_by_name,
                        user_ctx=user_ctx,
                        tool_roles=tool_roles,
                        conv_id=conv_id,
                        room_members=room_members,
                    ),
                    context=contextvars.copy_context(),
                )
                for tc in group
            ]
            # ``_run_one_tool`` catches its own exceptions and always
            # returns a ``ToolResult`` (error ones flagged ``is_error``),
            # so ``return_exceptions=False`` is correct — anything that
            # escapes here is a bug in the wrapper itself, not user code.
            gathered = await asyncio.gather(*tasks)
            for tr, blocks in gathered:
                results.append(tr)
                ui_blocks.extend(blocks)

        return results, ui_blocks

    @staticmethod
    def _group_is_parallel(
        group: list[ToolCall],
        tools_by_name: dict[str, tuple[ToolProvider, ToolDefinition]],
    ) -> bool:
        """Return True iff every call in ``group`` is parallel_safe."""
        for tc in group:
            pd = tools_by_name.get(tc.tool_name)
            if pd is None or not pd[1].parallel_safe:
                return False
        return True

    async def _run_one_tool(
        self,
        tc: ToolCall,
        tools_by_name: dict[str, tuple[ToolProvider, ToolDefinition]],
        *,
        user_ctx: UserContext | None,
        tool_roles: dict[str, str],
        conv_id: str | None,
        room_members: list[dict[str, str]] | None,
    ) -> tuple[ToolResult, list[UIBlock]]:
        """Execute a single tool call, handling RBAC, argument injection,
        lifecycle events, and exception capture.

        Returns ``(ToolResult, ui_blocks)``. Never raises — any error
        becomes a ``ToolResult`` with ``is_error=True`` so callers can
        keep iterating a batch without special-casing exceptions.
        """
        ui_blocks: list[UIBlock] = []

        provider_and_def = tools_by_name.get(tc.tool_name)
        if provider_and_def is None:
            return ToolResult(
                tool_call_id=tc.tool_call_id,
                content=f"Error: unknown tool '{tc.tool_name}'",
                is_error=True,
            ), ui_blocks
        provider, tool_def = provider_and_def

        # Defense in depth: re-check permission before execution.
        # Uses profile tool_roles overrides for consistency with _discover_tools.
        if user_ctx is not None and user_ctx.user_id != "system" and self._acl_svc is not None:
            if isinstance(self._acl_svc, AccessControlProvider):
                effective_role = tool_roles.get(tc.tool_name, tool_def.required_role)
                role_level = self._acl_svc.get_role_level(effective_role)
                user_level = self._acl_svc.get_effective_level(user_ctx)
                if user_level > role_level:
                    return ToolResult(
                        tool_call_id=tc.tool_call_id,
                        content=f"Permission denied: tool '{tc.tool_name}' requires higher privileges",
                        is_error=True,
                    ), ui_blocks

        # Per-task argument copy — parallel siblings must not share a
        # dict with each other or with ``ToolCall.arguments`` on the
        # caller's frozen ToolCall. The copy is shallow; tools that
        # mutate nested structures during execution should make their
        # own defensive copies of those, but the injected keys
        # themselves are scalar/list-typed and safe.
        arguments: dict[str, Any] = dict(tc.arguments)
        if user_ctx is not None:
            arguments["_user_id"] = user_ctx.user_id
            arguments["_user_name"] = user_ctx.display_name
            arguments["_user_roles"] = list(user_ctx.roles)
            if user_ctx.email:
                arguments["_user_email"] = user_ctx.email
        if conv_id:
            arguments["_conversation_id"] = conv_id
        arguments["_invocation_source"] = "ai"
        if room_members is not None:
            arguments["_room_members"] = room_members

        await self._publish_event(
            "chat.tool.started",
            {
                "conversation_id": conv_id,
                "tool_name": tc.tool_name,
                "tool_call_id": tc.tool_call_id,
                "arguments": self._sanitize_tool_args(arguments),
            },
        )

        # Propagate caller identity through the async context so tools
        # can resolve it via core.context.get_current_user(). When this
        # method runs inside a parallel asyncio.Task created with a
        # copied context, the .set() is local to that task and cannot
        # overwrite a sibling's user. On the serial singleton path, the
        # .set() persists in the running context just like it did before
        # this refactor, preserving the historical behavior.
        if user_ctx is not None:
            from gilbert.core.context import set_current_user

            set_current_user(user_ctx)

        try:
            raw_result = await provider.execute_tool(tc.tool_name, arguments)

            # Normalize: tools may return
            #   - ``str``                 (simple text)
            #   - ``ToolOutput``          (text + ui_blocks + attachments)
            #   - ``ToolResult``          (for callers that need full control)
            # The uniform internal shape is a ``ToolResult`` with an
            # ``attachments`` tuple that we collect at the turn level.
            # The tool's own ``tool_call_id`` is ignored — we rebind to
            # the live ``ToolCall.tool_call_id`` so the model sees a
            # result that matches its request id.
            tool_attachments: tuple[FileAttachment, ...] = ()
            result_is_error = False
            if isinstance(raw_result, ToolResult):
                result_text = raw_result.content
                tool_attachments = raw_result.attachments
                result_is_error = raw_result.is_error
            elif isinstance(raw_result, ToolOutput):
                result_text = raw_result.text
                tool_attachments = raw_result.attachments
                for block in raw_result.ui_blocks:
                    import dataclasses as _dc

                    # Auto-assign block_id if missing
                    if not block.block_id:
                        block = _dc.replace(block, block_id=str(uuid.uuid4()))
                    # Tag with tool name if not set
                    if not block.tool_name:
                        block = _dc.replace(block, tool_name=tc.tool_name)
                    ui_blocks.append(block)
            else:
                result_text = raw_result

            tool_result = ToolResult(
                tool_call_id=tc.tool_call_id,
                content=result_text,
                is_error=result_is_error,
                attachments=tool_attachments,
            )

            await self._publish_event(
                "chat.tool.completed",
                {
                    "conversation_id": conv_id,
                    "tool_name": tc.tool_name,
                    "tool_call_id": tc.tool_call_id,
                    "is_error": False,
                    "result_preview": result_text[:200] if result_text else "",
                },
            )
            return tool_result, ui_blocks
        except Exception as exc:
            logger.exception("Tool execution failed: %s", tc.tool_name)
            await self._publish_event(
                "chat.tool.completed",
                {
                    "conversation_id": conv_id,
                    "tool_name": tc.tool_name,
                    "tool_call_id": tc.tool_call_id,
                    "is_error": True,
                    "result_preview": str(exc)[:200],
                },
            )
            return ToolResult(
                tool_call_id=tc.tool_call_id,
                content=f"Error executing tool: {exc}",
                is_error=True,
            ), ui_blocks

    # --- Slash-command execution ---

    @staticmethod
    def _collapse_continuations(
        messages: list[Message],
        continuation_indices: set[int],
    ) -> list[Message]:
        """Remove synthetic max_tokens continuation user rows and merge the
        adjacent assistant text rows they used to split.

        After a max_tokens recovery sequence, the in-memory history looks
        like::

            USER (original)
            ASSISTANT (first chunk)
            USER  <-- synthetic "please continue"
            ASSISTANT (second chunk)
            USER  <-- synthetic "please continue"
            ASSISTANT (final chunk)

        which is valid for the backend (strict role alternation) but a
        bad thing to persist and show the user. After collapse::

            USER (original)
            ASSISTANT (first chunk\\n\\nsecond chunk\\n\\nfinal chunk)

        Assistant rows that carry ``tool_calls`` are never merged — those
        are part of the tool-use pairing with a following ``tool_result``
        row and must stay intact.

        ``continuation_indices`` is the set of positions in ``messages``
        that were added as synthetic user rows by the loop. Everything
        else is left alone, so the prior conversation history and any
        intermediate tool_result rows come through unchanged.
        """
        cleaned: list[Message] = [
            m for i, m in enumerate(messages) if i not in continuation_indices
        ]
        merged: list[Message] = []
        for msg in cleaned:
            if (
                merged
                and msg.role == MessageRole.ASSISTANT
                and merged[-1].role == MessageRole.ASSISTANT
                and not msg.tool_calls
                and not merged[-1].tool_calls
            ):
                prev = merged[-1]
                prev_text = prev.content or ""
                next_text = msg.content or ""
                if prev_text and next_text:
                    combined_content = f"{prev_text}\n\n{next_text}"
                elif prev_text:
                    combined_content = prev_text
                else:
                    combined_content = next_text
                merged[-1] = Message(
                    role=MessageRole.ASSISTANT,
                    content=combined_content,
                    tool_calls=[],
                    tool_results=list(prev.tool_results) + list(msg.tool_results),
                    author_id=prev.author_id or msg.author_id,
                    author_name=prev.author_name or msg.author_name,
                    visible_to=prev.visible_to,
                    attachments=list(prev.attachments) + list(msg.attachments),
                )
            else:
                merged.append(msg)
        return merged

    @staticmethod
    def _resolve_slash_namespace(provider: ToolProvider) -> str:
        """Figure out the slash-command namespace for *provider*, if any.

        Resolution order:

        1. If the provider's class declares ``slash_namespace`` as a
           non-empty string, use it verbatim. Plugins use this to pick a
           short human-friendly prefix (e.g. ``"currev"`` instead of
           ``"current-data-sync"``).
        2. If the provider's class was defined in a plugin module
           (``gilbert_plugin_<name>``), derive the namespace from the
           sanitized plugin name. This guarantees every plugin tool gets
           a namespace even if the plugin author forgets to set one.
        3. Otherwise (core service), return ``""`` — no prefix.
        """
        explicit = getattr(type(provider), "slash_namespace", "") or ""
        if explicit:
            return str(explicit)
        module = type(provider).__module__ or ""
        prefix = "gilbert_plugin_"
        if module.startswith(prefix):
            # ``gilbert_plugin_current_data_sync.data_sync_service`` →
            # ``current_data_sync``
            tail = module[len(prefix) :]
            return tail.split(".", 1)[0]
        return ""

    def _slash_commands_for_user(
        self,
        user_ctx: UserContext | None,
    ) -> dict[str, tuple[ToolProvider, ToolDefinition]]:
        """Return slash-enabled tools the user may invoke, keyed by full command name.

        Respects RBAC (via ``_discover_tools``) but ignores AI profile
        filtering — slash commands are user-initiated, not AI calls.

        The registry key is the full user-facing invocation string,
        reflecting both the plugin namespace (if any) and the tool's
        slash group (if any). Examples::

            "announce"                 # core, no group
            "radio start"              # core, group="radio", cmd="start"
            "currev.time_logs"         # plugin ns, no group
            "currev.sync status"       # plugin ns, group="sync", cmd="status"

        Plugin-sourced tools are automatically prefixed with their
        plugin namespace so they can't collide with core commands or
        with each other.
        """
        all_tools = self._discover_tools(user_ctx=user_ctx)
        result: dict[str, tuple[ToolProvider, ToolDefinition]] = {}
        for _tool_name, (provider, tool_def) in all_tools.items():
            cmd = tool_def.slash_command
            if not cmd:
                continue
            group = tool_def.slash_group
            local = f"{group} {cmd}" if group else cmd
            namespace = self._resolve_slash_namespace(provider)
            full_cmd = f"{namespace}.{local}" if namespace else local
            if full_cmd in result:
                logger.warning(
                    "Duplicate slash command %r from tool %r (already registered by %r)",
                    full_cmd,
                    tool_def.name,
                    result[full_cmd][1].name,
                )
                continue
            result[full_cmd] = (provider, tool_def)
        return result

    @staticmethod
    def _match_slash_command(
        text: str,
        registry: dict[str, tuple[ToolProvider, ToolDefinition]],
    ) -> str | None:
        """Longest-prefix lookup from an input line to a registered command.

        Given raw input like ``"/radio start some args"`` and a registry
        whose keys may include both grouped forms like ``"radio start"``
        and plain forms like ``"announce"``, return the longest matching
        key or ``None``.

        The algorithm tries the two-word form first (``"radio start"``)
        and falls back to the first-word form (``"radio"``). Plugin
        namespaces (``"currev.radio"`` / ``"currev.radio start"``) work
        because they use the first space as the separator between group
        and subcommand — the dot-prefixed namespace stays attached to
        the group.
        """
        stripped = text.lstrip()
        if not stripped.startswith("/"):
            return None
        body = stripped[1:]
        if not body:
            return None
        parts = body.split(None, 2)
        if not parts:
            return None
        first = parts[0]
        # Prefer the two-word (grouped) form when it matches.
        if len(parts) >= 2:
            candidate = f"{first} {parts[1]}"
            if candidate in registry:
                return candidate
        if first in registry:
            return first
        return None

    async def _execute_slash_command(
        self,
        raw_text: str,
        cmd_name: str,
        entry: tuple[ToolProvider, ToolDefinition],
        messages: list[Message],
        conversation_id: str,
        user_ctx: UserContext | None,
    ) -> ChatTurnResult:
        """Parse, execute, and persist a slash command.

        Returns the same ``ChatTurnResult`` shape as ``chat()`` so callers
        can't tell the difference between a slash command and an AI turn.
        """
        provider, tool_def = entry

        # Record the user's command as a user message (with author fields
        # so shared-room history renders the actor correctly).
        messages.append(
            Message(
                role=MessageRole.USER,
                content=raw_text,
                author_id=user_ctx.user_id if user_ctx else "",
                author_name=user_ctx.display_name if user_ctx else "",
            )
        )

        # Parse — errors are shown to the user as the assistant reply.
        # ``cmd_name`` is the matched full command (e.g. ``"radio start"``
        # or ``"currev.time_logs"``), passed explicitly so the parser
        # strips the correct prefix for grouped invocations.
        try:
            arguments = parse_slash_command(
                raw_text,
                tool_def,
                full_command=cmd_name,
            )
        except SlashCommandError as exc:
            error_text = str(exc)
            messages.append(
                Message(
                    role=MessageRole.ASSISTANT,
                    content=error_text,
                )
            )
            await self._save_conversation(
                conversation_id,
                messages,
                user_ctx=user_ctx,
            )
            return ChatTurnResult(
                response_text=error_text,
                conversation_id=conversation_id,
                ui_blocks=[],
                tool_usage=[
                    {
                        "tool_name": tool_def.name,
                        "is_error": True,
                        "arguments": {},
                        "result": error_text,
                    }
                ],
                attachments=[],
                rounds=[],
            )

        # Inject caller identity so tools can see who invoked them,
        # matching the AI-driven path in ``_execute_tool_calls``.
        if user_ctx is not None:
            arguments["_user_id"] = user_ctx.user_id
            arguments["_user_name"] = user_ctx.display_name
            arguments["_user_roles"] = list(user_ctx.roles)
            if user_ctx.email:
                arguments["_user_email"] = user_ctx.email

        # Slash-command source tag so SkillService (and any other tool
        # that wants to gate on activation) can let the call through —
        # slash commands are user-initiated, so they're a deliberate
        # "use this for this chat" signal.
        arguments["_conversation_id"] = conversation_id
        arguments["_invocation_source"] = "slash"

        # Inject shared-room members if this is a room conversation.
        if self._storage is not None:
            conv_data = await self._storage.get(_COLLECTION, conversation_id)
            if conv_data and conv_data.get("shared"):
                arguments["_room_members"] = [
                    {
                        "user_id": m.get("user_id", ""),
                        "display_name": m.get("display_name", ""),
                    }
                    for m in conv_data.get("members", [])
                ]

        tool_call_id = f"slash-{uuid.uuid4().hex[:12]}"
        sanitized_args = self._sanitize_tool_args(arguments)

        await self._publish_event(
            "chat.tool.started",
            {
                "conversation_id": conversation_id,
                "tool_name": tool_def.name,
                "tool_call_id": tool_call_id,
                "arguments": sanitized_args,
            },
        )

        # Propagate caller identity through the async context so
        # tools can resolve it via core.context.get_current_user().
        if user_ctx is not None:
            from gilbert.core.context import set_current_user

            set_current_user(user_ctx)

        ui_blocks: list[UIBlock] = []
        slash_attachments: list[FileAttachment] = []
        is_error = False
        try:
            raw_result = await provider.execute_tool(tool_def.name, arguments)
            if isinstance(raw_result, ToolResult):
                result_text = raw_result.content
                slash_attachments.extend(raw_result.attachments)
            elif isinstance(raw_result, ToolOutput):
                result_text = raw_result.text
                slash_attachments.extend(raw_result.attachments)
                import dataclasses as _dc

                for block in raw_result.ui_blocks:
                    if not block.block_id:
                        block = _dc.replace(block, block_id=str(uuid.uuid4()))
                    if not block.tool_name:
                        block = _dc.replace(block, tool_name=tool_def.name)
                    ui_blocks.append(block)
            else:
                result_text = raw_result
        except Exception as exc:
            logger.exception(
                "Slash command execution failed: /%s -> %s",
                cmd_name,
                tool_def.name,
            )
            result_text = f"Error executing /{cmd_name}: {exc}"
            is_error = True

        await self._publish_event(
            "chat.tool.completed",
            {
                "conversation_id": conversation_id,
                "tool_name": tool_def.name,
                "tool_call_id": tool_call_id,
                "is_error": is_error,
                "result_preview": result_text[:200] if result_text else "",
            },
        )

        # Store the assistant turn with ToolCall/ToolResult metadata so
        # the frontend renders it identically to an AI-driven tool use.
        #
        # If the tool produced UI blocks or attachments, those ARE the
        # user-visible output — showing the raw ``result_text`` (which
        # is often a JSON payload intended for the AI) as a chat bubble
        # just pollutes the transcript. Drop it from Message.content in
        # that case but keep it in ToolResult.content so later AI turns
        # can still read what happened.
        if ui_blocks or slash_attachments:
            bubble_content = ""
        else:
            bubble_content = result_text
        messages.append(
            Message(
                role=MessageRole.ASSISTANT,
                content=bubble_content,
                tool_calls=[
                    ToolCall(
                        tool_call_id=tool_call_id,
                        tool_name=tool_def.name,
                        arguments=sanitized_args,
                    )
                ],
                tool_results=[
                    ToolResult(
                        tool_call_id=tool_call_id,
                        content=result_text,
                        is_error=is_error,
                        attachments=tuple(slash_attachments),
                    )
                ],
                attachments=list(slash_attachments),
            )
        )

        # Serialize UI blocks with position + submission state, matching
        # the chat() agentic loop so downstream rendering is uniform.
        # ``response_index`` anchors UI blocks to the turn they were
        # produced in. The frontend renders one ``TurnBubble`` per
        # user→assistant exchange (one turn per user message), so we
        # count user rows in the persisted history and use that as the
        # turn index. This matches ``MessageList``'s
        # ``assistantToTurnIndex`` map exactly — both sides agree on
        # "the Nth turn is the Nth user message."
        user_count = sum(1 for m in messages if m.role == MessageRole.USER)
        response_index = max(0, user_count - 1)
        ui_block_dicts: list[dict[str, Any]] = []
        for block in ui_blocks:
            d = block.to_dict()
            d["response_index"] = response_index
            d["submitted"] = False
            d["submission"] = None
            ui_block_dicts.append(d)

        await self._save_conversation(
            conversation_id,
            messages,
            user_ctx=user_ctx,
            ui_blocks=ui_block_dicts,
        )

        tool_usage = [
            {
                "tool_name": tool_def.name,
                "is_error": is_error,
                "arguments": sanitized_args,
                "result": result_text,
            }
        ]
        # Slash commands are modeled as a single-round turn so the
        # frontend's TurnBubble shows the "thinking" card with the tool
        # call + result inside it — mirroring what refreshing the page
        # would reconstruct from persisted history. If the tool produced
        # UI blocks or attachments, those are the visible output and the
        # bubble's final text is left empty (matches
        # ``bubble_content`` above). Otherwise, ``result_text`` is the
        # user-facing answer.
        synthetic_round = {
            "reasoning": "",
            "tools": [
                {
                    "tool_call_id": tool_call_id,
                    "tool_name": tool_def.name,
                    "is_error": is_error,
                    "arguments": sanitized_args,
                    "result": result_text,
                }
            ],
        }
        if ui_blocks or slash_attachments:
            live_response_text = ""
        else:
            live_response_text = result_text
        return ChatTurnResult(
            response_text=live_response_text,
            conversation_id=conversation_id,
            ui_blocks=ui_block_dicts,
            tool_usage=tool_usage,
            attachments=list(slash_attachments),
            rounds=[synthetic_round],
        )

    # --- Tool Event Publishing ---

    async def _resolve_stream_audience(
        self,
        conversation_id: str,
        user_ctx: UserContext | None,
    ) -> list[str]:
        """List the user_ids that should receive live streaming events.

        Computed once at the start of a chat turn and attached as
        ``visible_to`` on ``chat.stream.*`` events so the WS layer can
        deliver them only to the right connections (see the
        ``chat.stream.`` branch in ``WsConnection.can_see_chat_event``).

        Resolution rules:

        - Personal conversation: owner only. In a clean DB this is just
          ``[owner_id]``; when the conversation doesn't exist yet
          (new-turn slash path), fall back to the caller's user_id.
        - Shared room: every current member's user_id.

        System callers (scheduler, greeting, roast, …) have no browser
        connection to stream to, so the audience is empty — events are
        still published but no WS client sees them.
        """
        fallback: list[str] = []
        if user_ctx is not None and user_ctx.user_id not in ("", "system"):
            fallback = [user_ctx.user_id]
        if self._storage is None:
            return fallback
        conv = await self._storage.get(_COLLECTION, conversation_id)
        if not isinstance(conv, dict):
            return fallback
        if conv.get("shared"):
            members = conv.get("members") or []
            ids = [
                str(m.get("user_id", ""))
                for m in members
                if isinstance(m, dict) and m.get("user_id")
            ]
            return ids or fallback
        owner = conv.get("user_id")
        if owner:
            return [str(owner)]
        return fallback

    async def _publish_event(
        self,
        event_type: str,
        data: dict[str, Any],
    ) -> None:
        """Publish an event for real-time UI updates.

        Thin wrapper around the event bus — the ``ai`` source tag lets
        downstream consumers (telemetry, audit logging, peer forwarding)
        filter by origin. Used for tool lifecycle events
        (``chat.tool.started`` / ``chat.tool.completed``) and streaming
        text deltas (``chat.message.text_delta``). No-op when there is
        no resolver or no event bus service available.
        """
        if self._resolver is None:
            return
        event_bus_svc = self._resolver.get_capability("event_bus")
        if event_bus_svc is None:
            return
        from gilbert.interfaces.events import Event, EventBusProvider

        if isinstance(event_bus_svc, EventBusProvider):
            await event_bus_svc.bus.publish(
                Event(
                    event_type=event_type,
                    data=data,
                    source="ai",
                )
            )

    @staticmethod
    def _sanitize_tool_args(args: dict[str, Any]) -> dict[str, Any]:
        """Remove injected internal arguments before sending to frontend."""
        return {k: v for k, v in args.items() if not k.startswith("_")}

    # --- Conversation Persistence ---

    async def _save_conversation(
        self,
        conv_id: str,
        messages: list[Message],
        user_ctx: UserContext | None = None,
        ui_blocks: list[dict[str, Any]] | None = None,
    ) -> None:
        """Persist a conversation to storage with optional user ownership."""
        if self._storage is None:
            return
        # Load existing data to preserve fields like title
        existing = await self._storage.get(_COLLECTION, conv_id) or {}
        data: dict[str, Any] = {
            **existing,
            "messages": [self._serialize_message(m) for m in messages],
            "updated_at": datetime.now(UTC).isoformat(),
        }
        if user_ctx is not None and user_ctx.user_id != "system":
            data["user_id"] = user_ctx.user_id

        # Merge new UI blocks with any existing ones
        if ui_blocks:
            existing_blocks: list[dict[str, Any]] = data.get("ui_blocks", [])
            existing_blocks.extend(ui_blocks)
            data["ui_blocks"] = existing_blocks

        await self._storage.put(_COLLECTION, conv_id, data)

    async def list_conversations(
        self, user_id: str | None = None, limit: int = 50
    ) -> list[dict[str, Any]]:
        """List personal (non-shared) conversations, optionally filtered by owning user."""
        if self._storage is None:
            return []
        filters: list[Filter] = []
        if user_id:
            filters.append(Filter(field="user_id", op=FilterOp.EQ, value=user_id))
        results = await self._storage.query(
            Query(
                collection=_COLLECTION,
                filters=filters,
                sort=[SortField(field="updated_at", descending=True)],
                limit=limit * 2,  # fetch extra to account for shared filtering
            )
        )
        # Exclude shared conversations — those are listed separately.
        # Can't use NEQ filter because shared=None (missing field) doesn't match.
        return [c for c in results if not c.get("shared")][:limit]

    async def list_shared_conversations(
        self, user_id: str, limit: int = 50
    ) -> list[dict[str, Any]]:
        """List shared conversations visible to user_id.

        Returns conversations where the user is a member, plus public rooms
        they haven't joined yet (so they can see and join them).
        """
        if self._storage is None:
            return []
        shared = await self._storage.query(
            Query(
                collection=_COLLECTION,
                filters=[Filter(field="shared", op=FilterOp.EQ, value=True)],
                sort=[SortField(field="updated_at", descending=True)],
                limit=200,
            )
        )
        results = []
        for conv in shared:
            members = conv.get("members", [])
            invites = conv.get("invites", [])
            is_member = any(m.get("user_id") == user_id for m in members)
            is_invited = any(inv.get("user_id") == user_id for inv in invites)
            is_public = conv.get("visibility") == "public"
            if is_member or is_invited or is_public:
                conv["_is_member"] = is_member
                conv["_is_invited"] = is_invited
                results.append(conv)
                if len(results) >= limit:
                    break
        return results

    async def _load_conversation(self, conv_id: str) -> list[Message]:
        """Load a conversation from storage. Returns empty list if not found."""
        if self._storage is None:
            return []
        data = await self._storage.get(_COLLECTION, conv_id)
        if data is None:
            return []
        return [self._deserialize_message(m) for m in data.get("messages", [])]

    @staticmethod
    def _serialize_message(msg: Message) -> dict[str, Any]:
        d: dict[str, Any] = {"role": msg.role.value, "content": msg.content}
        if msg.tool_calls:
            d["tool_calls"] = [
                {
                    "tool_call_id": tc.tool_call_id,
                    "tool_name": tc.tool_name,
                    "arguments": tc.arguments,
                }
                for tc in msg.tool_calls
            ]
        if msg.tool_results:
            d["tool_results"] = [
                {
                    "tool_call_id": tr.tool_call_id,
                    "content": tr.content,
                    "is_error": tr.is_error,
                }
                for tr in msg.tool_results
            ]
        if msg.author_id:
            d["author_id"] = msg.author_id
        if msg.author_name:
            d["author_name"] = msg.author_name
        if msg.visible_to is not None:
            d["visible_to"] = msg.visible_to
        if msg.attachments:
            serialized_attachments: list[dict[str, Any]] = []
            for att in msg.attachments:
                entry: dict[str, Any] = {
                    "kind": att.kind,
                    "name": att.name,
                    "media_type": att.media_type,
                }
                if att.data:
                    entry["data"] = att.data
                if att.text:
                    entry["text"] = att.text
                # Reference-mode attachments: persist the workspace
                # coordinates so the frontend can fetch bytes on click via
                # ``skills.workspace.download``.
                if att.workspace_skill:
                    entry["workspace_skill"] = att.workspace_skill
                if att.workspace_path:
                    entry["workspace_path"] = att.workspace_path
                if att.workspace_conv:
                    entry["workspace_conv"] = att.workspace_conv
                if att.workspace_file_id:
                    entry["workspace_file_id"] = att.workspace_file_id
                # ``size`` is meaningful for reference-mode files
                # (where there are no bytes in the row to count) but
                # also useful for inline kinds so history loads don't
                # have to redecode base64 to show a size chip.
                if att.size:
                    entry["size"] = att.size
                serialized_attachments.append(entry)
            d["attachments"] = serialized_attachments
        if msg.interrupted:
            d["interrupted"] = True
        if msg.usage:
            d["usage"] = msg.usage
        return d

    @staticmethod
    def _deserialize_message(data: dict[str, Any]) -> Message:
        tool_calls = [
            ToolCall(
                tool_call_id=tc["tool_call_id"],
                tool_name=tc["tool_name"],
                arguments=tc["arguments"],
            )
            for tc in data.get("tool_calls", [])
        ]
        tool_results = [
            ToolResult(
                tool_call_id=tr["tool_call_id"],
                content=tr["content"],
                is_error=tr.get("is_error", False),
            )
            for tr in data.get("tool_results", [])
        ]
        attachments: list[FileAttachment] = []
        raw_attachments = data.get("attachments")
        if isinstance(raw_attachments, list):
            for att in raw_attachments:
                if not isinstance(att, dict):
                    continue
                kind = str(att.get("kind") or "")
                if not kind:
                    continue
                raw_size = att.get("size")
                try:
                    size_val = int(raw_size) if raw_size is not None else 0
                except (TypeError, ValueError):
                    size_val = 0
                attachments.append(
                    FileAttachment(
                        kind=kind,
                        name=str(att.get("name", "")),
                        media_type=str(att.get("media_type", "")),
                        data=str(att.get("data", "")),
                        text=str(att.get("text", "")),
                        workspace_skill=str(att.get("workspace_skill", "")),
                        workspace_path=str(att.get("workspace_path", "")),
                        workspace_conv=str(att.get("workspace_conv", "")),
                        workspace_file_id=str(att.get("workspace_file_id", "")),
                        size=size_val,
                    )
                )
        else:
            # Legacy: pre-attachments schema stored images under "images".
            for img in data.get("images", []) or []:
                if isinstance(img, dict) and img.get("data"):
                    attachments.append(
                        FileAttachment(
                            kind="image",
                            media_type=str(img.get("media_type", "")),
                            data=str(img.get("data", "")),
                        )
                    )
        raw_usage = data.get("usage")
        usage = raw_usage if isinstance(raw_usage, dict) else None
        return Message(
            role=MessageRole(data["role"]),
            content=data.get("content", ""),
            tool_calls=tool_calls,
            tool_results=tool_results,
            author_id=data.get("author_id", ""),
            author_name=data.get("author_name", ""),
            visible_to=data.get("visible_to"),
            attachments=attachments,
            interrupted=bool(data.get("interrupted", False)),
            usage=usage,
        )

    # --- Conversation State ---

    def _resolve_conversation_id(self, conversation_id: str | None) -> str:
        """Resolve to an explicit or the current conversation ID."""
        cid = conversation_id or get_current_conversation_id()
        if not cid:
            raise RuntimeError("No active conversation")
        return cid

    async def get_conversation_state(
        self,
        key: str,
        conversation_id: str | None = None,
    ) -> Any | None:
        """Read a state entry from a conversation.

        Args:
            key: Namespace key (e.g. ``"guess_game"``).
            conversation_id: Explicit conversation ID, or ``None`` to use the
                currently active conversation.

        Returns:
            The stored value, or ``None`` if the key doesn't exist.
        """
        if self._storage is None:
            return None
        cid = self._resolve_conversation_id(conversation_id)
        data = await self._storage.get(_COLLECTION, cid)
        if data is None:
            return None
        return data.get("state", {}).get(key)

    async def set_conversation_state(
        self,
        key: str,
        value: Any,
        conversation_id: str | None = None,
    ) -> None:
        """Write a state entry to a conversation.

        The value must be JSON-serialisable.  It is persisted immediately so
        that subsequent agentic-loop rounds see the update.

        Args:
            key: Namespace key (e.g. ``"guess_game"``).
            value: Any JSON-serialisable value.
            conversation_id: Explicit conversation ID, or ``None`` to use the
                currently active conversation.
        """
        if self._storage is None:
            return
        cid = self._resolve_conversation_id(conversation_id)
        data = await self._storage.get(_COLLECTION, cid) or {}
        state: dict[str, Any] = data.get("state", {})
        state[key] = value
        data["state"] = state
        await self._storage.put(_COLLECTION, cid, data)

    async def clear_conversation_state(
        self,
        key: str,
        conversation_id: str | None = None,
    ) -> None:
        """Remove a state entry from a conversation.

        Args:
            key: Namespace key to remove.
            conversation_id: Explicit conversation ID, or ``None`` to use the
                currently active conversation.
        """
        if self._storage is None:
            return
        cid = self._resolve_conversation_id(conversation_id)
        data = await self._storage.get(_COLLECTION, cid)
        if data is None:
            return
        state: dict[str, Any] = data.get("state", {})
        if key in state:
            del state[key]
            data["state"] = state
            await self._storage.put(_COLLECTION, cid, data)

    async def _load_conversation_state(self, conv_id: str) -> dict[str, Any]:
        """Load all state entries for a conversation."""
        if self._storage is None:
            return {}
        data = await self._storage.get(_COLLECTION, conv_id)
        if data is None:
            return {}
        state = data.get("state", {})
        return state if isinstance(state, dict) else {}

    @staticmethod
    def _format_state_for_context(state: dict[str, Any]) -> str:
        """Render conversation state as a text block for the system prompt."""
        parts: list[str] = ["## Active Conversation State"]
        for key, value in state.items():
            parts.append(f"\n### {key}")
            if isinstance(value, (dict, list)):
                parts.append(_json.dumps(value, indent=2, default=str))
            else:
                parts.append(str(value))
        return "\n".join(parts)

    # --- History Management ---

    def _get_effective_compression_config(
        self, per_conv: dict[str, Any] | None
    ) -> dict[str, Any]:
        """Merge per-conversation compression overrides with global defaults."""
        defaults = {
            "enabled": self._compression_enabled,
            "threshold": self._compression_threshold,
            "keep_recent": self._compression_keep_recent,
            "summary_max_tokens": self._compression_summary_max_tokens,
        }
        if per_conv:
            defaults.update({k: v for k, v in per_conv.items() if v is not None})
        return defaults

    @staticmethod
    def _find_clean_boundary(messages: list[Message], raw_idx: int) -> int:
        """Adjust a split index so it doesn't land between a tool_call and its result."""
        idx = raw_idx
        while idx < len(messages) and messages[idx].role == MessageRole.TOOL_RESULT:
            idx += 1
        return min(idx, len(messages))

    async def _maybe_compress_history(
        self,
        messages: list[Message],
        conversation_id: str,
        *,
        force: bool = False,
    ) -> None:
        """Summarize older messages if the conversation exceeds the threshold.

        The summary is persisted in conversation state and injected into the
        system prompt on subsequent turns. The original messages are kept in
        storage — compression only affects what gets sent to the backend.

        When *force* is True, the threshold check is skipped (but
        ``keep_recent`` is still respected).
        """
        if not self._backends:
            return

        per_conv_config = await self.get_conversation_state(
            _COMPRESSION_CONFIG_KEY, conversation_id
        )
        cfg = self._get_effective_compression_config(
            per_conv_config if isinstance(per_conv_config, dict) else None
        )

        if not force:
            if not cfg["enabled"]:
                return
            if len(messages) <= cfg["threshold"]:
                return

        existing = await self.get_conversation_state(
            _COMPRESSION_STATE_KEY, conversation_id
        )
        compressed_up_to = (
            existing.get("compressed_up_to", 0)
            if isinstance(existing, dict)
            else 0
        )
        existing_summary = (
            existing.get("summary", "") if isinstance(existing, dict) else ""
        )

        keep_recent: int = cfg["keep_recent"]
        new_boundary = self._find_clean_boundary(
            messages, max(len(messages) - keep_recent, compressed_up_to)
        )

        if new_boundary <= compressed_up_to:
            return

        chunk = messages[compressed_up_to:new_boundary]
        if not chunk:
            return

        chunk_text_parts: list[str] = []
        if existing_summary:
            chunk_text_parts.append(f"EXISTING SUMMARY:\n{existing_summary}\n")
        chunk_text_parts.append("NEW MESSAGES TO INCORPORATE:")
        for msg in chunk:
            role_label = msg.role.value.upper()
            line = f"[{role_label}] {msg.content}"
            if msg.attachments:
                att_descs = []
                for att in msg.attachments:
                    desc = att.name or "unnamed"
                    if att.media_type:
                        desc += f" ({att.media_type})"
                    att_descs.append(desc)
                line += f" [attachments: {', '.join(att_descs)}]"
            if msg.tool_calls:
                tool_names = ", ".join(tc.tool_name for tc in msg.tool_calls)
                line += f" [called: {tool_names}]"
            if msg.tool_results:
                for tr in msg.tool_results:
                    snippet = tr.content[:200] if tr.content else ""
                    err = " (ERROR)" if tr.is_error else ""
                    line += f"\n  -> result{err}: {snippet}"
            chunk_text_parts.append(line)

        await self._publish_event(
            "chat.compression.started",
            {
                "conversation_id": conversation_id,
                "messages_to_compress": len(chunk),
                "total_messages": len(messages),
            },
        )

        summary_request = AIRequest(
            messages=[
                Message(
                    role=MessageRole.USER,
                    content="\n".join(chunk_text_parts),
                )
            ],
            system_prompt=_COMPRESSION_SYSTEM_PROMPT,
        )

        try:
            resp = await self.get_backend().generate(summary_request)
            summary_text = resp.message.content.strip()
        except Exception:
            logger.warning(
                "Compression summarization failed for conversation %s; "
                "falling back to plain truncation",
                conversation_id,
                exc_info=True,
            )
            await self._publish_event(
                "chat.compression.failed",
                {"conversation_id": conversation_id},
            )
            return

        await self.set_conversation_state(
            _COMPRESSION_STATE_KEY,
            {
                "summary": summary_text,
                "compressed_up_to": new_boundary,
                "last_compressed_at": datetime.now(UTC).isoformat(),
            },
            conversation_id,
        )

        await self._publish_event(
            "chat.compression.completed",
            {
                "conversation_id": conversation_id,
                "messages_compressed": len(chunk),
                "compressed_up_to": new_boundary,
            },
        )
        logger.info(
            "Compressed %d messages in conversation %s (up to index %d)",
            len(chunk),
            conversation_id,
            new_boundary,
        )

    def _truncate_history(
        self,
        messages: list[Message],
        compression_state: dict[str, Any] | None = None,
    ) -> list[Message]:
        """Truncate to max_history_messages, preserving tool-call/result pairs.

        When a compression summary exists, slices from ``compressed_up_to``
        instead of a blind last-N. The summary itself is injected into the
        system prompt separately — this method only decides which raw messages
        to include in the request.
        """
        start = 0
        if compression_state and isinstance(compression_state.get("compressed_up_to"), int):
            start = compression_state["compressed_up_to"]

        tail = messages[start:]

        if len(tail) > self._max_history_messages:
            tail = tail[-self._max_history_messages :]

        while tail and tail[0].role == MessageRole.TOOL_RESULT:
            idx = messages.index(tail[0])
            if idx > 0:
                tail.insert(0, messages[idx - 1])
            else:
                break

        return tail

    # --- ToolProvider protocol ---

    @property
    def tool_provider_name(self) -> str:
        return "ai"

    def get_tools(self, user_ctx: UserContext | None = None) -> list[ToolDefinition]:
        if not self._enabled:
            return []
        tools = [
            ToolDefinition(
                name="rename_conversation",
                slash_command="rename",
                slash_help="Rename the current conversation: /rename <title>",
                description="Rename the current chat conversation to a user-specified title.",
                parameters=[
                    ToolParameter(
                        name="title",
                        type=ToolParameterType.STRING,
                        description="The new title for this conversation.",
                    ),
                ],
                required_role="everyone",
            ),
            ToolDefinition(
                name="list_ai_profiles",
                slash_group="profile",
                slash_command="list",
                slash_help="List AI profiles and call assignments: /profile list",
                description="List all AI context profiles and their call assignments.",
                required_role="admin",
            ),
            ToolDefinition(
                name="set_ai_profile",
                description=(
                    "Create or update an AI context profile. "
                    "tool_mode: 'all' (every tool), 'include' (only listed), 'exclude' (all except listed). "
                    "tool_roles: per-tool role overrides within this profile."
                ),
                parameters=[
                    ToolParameter(
                        name="name", type=ToolParameterType.STRING, description="Profile name."
                    ),
                    ToolParameter(
                        name="description",
                        type=ToolParameterType.STRING,
                        description="What this profile is for.",
                        required=False,
                    ),
                    ToolParameter(
                        name="tool_mode",
                        type=ToolParameterType.STRING,
                        description="'all', 'include', or 'exclude'.",
                        required=False,
                        enum=["all", "include", "exclude"],
                    ),
                    ToolParameter(
                        name="tools",
                        type=ToolParameterType.ARRAY,
                        description="Tool names for include/exclude mode.",
                        required=False,
                    ),
                    ToolParameter(
                        name="tool_roles",
                        type=ToolParameterType.OBJECT,
                        description="Per-tool role overrides: {tool_name: role_name}.",
                        required=False,
                    ),
                    ToolParameter(
                        name="backend",
                        type=ToolParameterType.STRING,
                        description="Backend name (e.g. 'anthropic'). Empty = first available.",
                        required=False,
                    ),
                    ToolParameter(
                        name="model",
                        type=ToolParameterType.STRING,
                        description="Model ID. Empty = backend's default.",
                        required=False,
                    ),
                ],
                required_role="admin",
            ),
            ToolDefinition(
                name="delete_ai_profile",
                slash_group="profile",
                slash_command="delete",
                slash_help="Delete an AI profile: /profile delete <name>",
                description="Delete an AI context profile. The 'default' profile cannot be deleted.",
                parameters=[
                    ToolParameter(
                        name="name",
                        type=ToolParameterType.STRING,
                        description="Profile name to delete.",
                    ),
                ],
                required_role="admin",
            ),
            ToolDefinition(
                name="assign_ai_profile",
                slash_group="profile",
                slash_command="assign",
                slash_help=(
                    "Assign a profile to an AI call name: /profile assign <call_name> <profile>"
                ),
                description="Assign an AI context profile to a named AI call (e.g., 'human_chat', 'sales_initial_email').",
                parameters=[
                    ToolParameter(
                        name="call_name",
                        type=ToolParameterType.STRING,
                        description="The AI call name.",
                    ),
                    ToolParameter(
                        name="profile",
                        type=ToolParameterType.STRING,
                        description="Profile name to assign.",
                    ),
                ],
                required_role="admin",
            ),
            ToolDefinition(
                name="clear_ai_assignment",
                slash_group="profile",
                slash_command="unassign",
                slash_help=(
                    "Revert a call to the 'default' profile: /profile unassign <call_name>"
                ),
                description="Remove a call's profile assignment, reverting it to the 'default' profile.",
                parameters=[
                    ToolParameter(
                        name="call_name",
                        type=ToolParameterType.STRING,
                        description="The AI call name.",
                    ),
                ],
                required_role="admin",
            ),
            # Persona tools
            ToolDefinition(
                name="get_persona",
                slash_group="persona",
                slash_command="show",
                slash_help="Show the current AI persona: /persona show",
                description="Get the current AI persona (personality, tone, and behavioral instructions).",
                required_role="everyone",
            ),
            ToolDefinition(
                name="update_persona",
                description=(
                    "Update the AI persona. This changes how Gilbert behaves, speaks, "
                    "and responds. The full persona text is replaced."
                ),
                parameters=[
                    ToolParameter(
                        name="text",
                        type=ToolParameterType.STRING,
                        description="The new persona text (full replacement).",
                    ),
                ],
                required_role="admin",
                # No slash_command: persona text is typically multi-line
                # (paragraphs of behavioral instructions); inline shell
                # quoting is impractical. Edit persona from the chat sidebar in the UI.
            ),
            ToolDefinition(
                name="reset_persona",
                slash_group="persona",
                slash_command="reset",
                slash_help="Reset persona to the default: /persona reset",
                description="Reset the AI persona to the default.",
                required_role="admin",
            ),
        ]
        # Memory tool (only when enabled)
        if self._memory_enabled:
            tools.append(
                ToolDefinition(
                    name="memory",
                    slash_command="memory",
                    slash_help=(
                        "Manage memories: /memory <action> "
                        "[summary='...'] [content='...'] "
                        "(actions: remember, recall, update, forget, list)"
                    ),
                    description=(
                        "Manage persistent memories for the current user. "
                        "Use 'remember' when the user tells you something worth remembering "
                        "(preferences, project details, personal info). Use 'auto' source when "
                        "you notice something worth remembering that the user didn't explicitly ask to save. "
                        "Use 'list' to see what you remember about them. "
                        "Use 'recall' to load full content of specific memories by ID. "
                        "Use 'update' to modify a memory. Use 'forget' to delete one."
                    ),
                    parameters=[
                        ToolParameter(
                            name="action",
                            type=ToolParameterType.STRING,
                            description="Action to perform.",
                            enum=["remember", "recall", "update", "forget", "list"],
                        ),
                        ToolParameter(
                            name="summary",
                            type=ToolParameterType.STRING,
                            description="Short summary sentence (for remember, or update).",
                            required=False,
                        ),
                        ToolParameter(
                            name="content",
                            type=ToolParameterType.STRING,
                            description="Detailed memory content (for remember, or update).",
                            required=False,
                        ),
                        ToolParameter(
                            name="source",
                            type=ToolParameterType.STRING,
                            description="'user' if they explicitly asked to remember, 'auto' if you decided to.",
                            enum=["user", "auto"],
                            required=False,
                        ),
                        ToolParameter(
                            name="ids",
                            type=ToolParameterType.ARRAY,
                            description="Memory IDs to recall (for recall action).",
                            required=False,
                        ),
                        ToolParameter(
                            name="id",
                            type=ToolParameterType.STRING,
                            description="Memory ID (for update or forget).",
                            required=False,
                        ),
                    ],
                    required_role="user",
                ),
            )
        tools.extend(
            [
                ToolDefinition(
                    name="get_compression_config",
                    slash_group="compression",
                    slash_command="show",
                    slash_help="Show compression settings for this conversation",
                    description=(
                        "Show the effective compression settings and current "
                        "compression state for the active conversation."
                    ),
                    required_role="everyone",
                ),
                ToolDefinition(
                    name="set_compression_config",
                    slash_group="compression",
                    slash_command="set",
                    slash_help=(
                        "Override compression settings: "
                        "/compression set [enabled] [threshold] [keep_recent] [summary_max_tokens]"
                    ),
                    description=(
                        "Set per-conversation compression overrides. Only "
                        "provided fields are updated; omitted fields keep "
                        "their current value."
                    ),
                    parameters=[
                        ToolParameter(
                            name="enabled",
                            type=ToolParameterType.BOOLEAN,
                            description="Enable or disable compression for this conversation.",
                            required=False,
                        ),
                        ToolParameter(
                            name="threshold",
                            type=ToolParameterType.INTEGER,
                            description="Compress when message count exceeds this.",
                            required=False,
                        ),
                        ToolParameter(
                            name="keep_recent",
                            type=ToolParameterType.INTEGER,
                            description="Number of recent messages to always keep verbatim.",
                            required=False,
                        ),
                        ToolParameter(
                            name="summary_max_tokens",
                            type=ToolParameterType.INTEGER,
                            description="Max tokens for the generated summary.",
                            required=False,
                        ),
                    ],
                    required_role="user",
                ),
                ToolDefinition(
                    name="clear_compression",
                    slash_group="compression",
                    slash_command="clear",
                    slash_help="Clear compression state (summary) for this conversation",
                    description=(
                        "Clear the compression summary for the active "
                        "conversation. The full message history is preserved; "
                        "a new summary will be generated when the threshold "
                        "is exceeded again."
                    ),
                    required_role="user",
                ),
                ToolDefinition(
                    name="force_compression",
                    slash_group="compression",
                    slash_command="compress",
                    slash_help="Force compression now, ignoring the threshold",
                    description=(
                        "Force-compress the current conversation immediately, "
                        "regardless of message count. Useful for long "
                        "conversations that haven't hit the threshold yet."
                    ),
                    required_role="user",
                ),
            ]
        )
        return tools

    async def execute_tool(self, name: str, arguments: dict[str, Any]) -> str:
        match name:
            case "rename_conversation":
                return await self._tool_rename_conversation(arguments)
            case "list_ai_profiles":
                return self._tool_list_profiles()
            case "set_ai_profile":
                return await self._tool_set_profile(arguments)
            case "delete_ai_profile":
                return await self._tool_delete_profile(arguments)
            case "assign_ai_profile":
                return await self._tool_assign_profile(arguments)
            case "clear_ai_assignment":
                return await self._tool_clear_assignment(arguments)
            case "get_persona":
                return await self._tool_get_persona()
            case "update_persona":
                return await self._tool_update_persona(arguments)
            case "reset_persona":
                return await self._tool_reset_persona()
            case "memory":
                return await self._tool_memory_action(arguments)
            case "get_compression_config":
                return await self._tool_get_compression_config(arguments)
            case "set_compression_config":
                return await self._tool_set_compression_config(arguments)
            case "clear_compression":
                return await self._tool_clear_compression(arguments)
            case "force_compression":
                return await self._tool_force_compression(arguments)
            case _:
                raise KeyError(f"Unknown tool: {name}")

    def _tool_list_profiles(self) -> str:
        profiles = []
        for p in self.list_profiles():
            profiles.append(
                {
                    "name": p.name,
                    "description": p.description,
                    "tool_mode": p.tool_mode,
                    "tools": p.tools,
                    "tool_roles": p.tool_roles,
                    "backend": p.backend,
                    "model": p.model,
                }
            )
        return _json.dumps(
            {
                "profiles": profiles,
                "assignments": self.list_assignments(),
            }
        )

    async def _tool_set_profile(self, arguments: dict[str, Any]) -> str:
        name = arguments.get("name", "").strip()
        if not name:
            return _json.dumps({"error": "Profile name is required"})
        existing = self._profiles.get(name)
        profile = AIContextProfile(
            name=name,
            description=arguments.get("description", existing.description if existing else ""),
            tool_mode=arguments.get("tool_mode", existing.tool_mode if existing else "all"),
            tools=arguments.get("tools", existing.tools if existing else []),
            tool_roles=arguments.get("tool_roles", existing.tool_roles if existing else {}),
            backend=arguments.get("backend", existing.backend if existing else ""),
            model=arguments.get("model", existing.model if existing else ""),
        )
        await self.set_profile(profile)
        return _json.dumps({"status": "saved", "profile": name})

    async def _tool_delete_profile(self, arguments: dict[str, Any]) -> str:
        try:
            await self.delete_profile(arguments["name"])
            return _json.dumps({"status": "deleted"})
        except (KeyError, ValueError) as e:
            return _json.dumps({"error": str(e)})

    async def _tool_assign_profile(self, arguments: dict[str, Any]) -> str:
        try:
            await self.set_assignment(arguments["call_name"], arguments["profile"])
            return _json.dumps({"status": "assigned"})
        except ValueError as e:
            return _json.dumps({"error": str(e)})

    async def _tool_clear_assignment(self, arguments: dict[str, Any]) -> str:
        await self.clear_assignment(arguments["call_name"])
        return _json.dumps({"status": "cleared"})

    # --- Compression tool handlers ---

    async def _tool_get_compression_config(self, arguments: dict[str, Any]) -> str:
        conv_id = self._resolve_conversation_id(arguments.get("_conversation_id"))
        per_conv = await self.get_conversation_state(_COMPRESSION_CONFIG_KEY, conv_id)
        effective = self._get_effective_compression_config(
            per_conv if isinstance(per_conv, dict) else None
        )
        state = await self.get_conversation_state(_COMPRESSION_STATE_KEY, conv_id)
        messages = await self._load_conversation(conv_id)

        lines = ["**Compression Settings**"]
        source = "per-conversation override" if per_conv else "global defaults"
        lines.append(f"Source: {source}")
        lines.append(f"Enabled: {'yes' if effective['enabled'] else 'no'}")
        lines.append(f"Threshold: {effective['threshold']} messages")
        lines.append(f"Keep recent: {effective['keep_recent']} messages")
        lines.append(f"Summary max tokens: {effective['summary_max_tokens']}")
        lines.append(f"Total messages in conversation: {len(messages)}")

        if isinstance(state, dict) and state.get("summary"):
            lines.append("")
            lines.append("**Current Compression State**")
            lines.append(
                f"Messages summarized: {state.get('compressed_up_to', 0)} "
                f"of {len(messages)}"
            )
            lines.append(
                f"Summary length: {len(state['summary']):,} characters"
            )
            ts = state.get("last_compressed_at", "")
            if ts:
                lines.append(f"Last compressed: {ts}")
        else:
            lines.append("")
            lines.append("No compression has been performed yet.")

        return "\n".join(lines)

    async def _tool_set_compression_config(self, arguments: dict[str, Any]) -> str:
        conv_id = self._resolve_conversation_id(arguments.get("_conversation_id"))
        existing = await self.get_conversation_state(_COMPRESSION_CONFIG_KEY, conv_id)
        merged: dict[str, Any] = existing if isinstance(existing, dict) else {}
        changed: list[str] = []
        for key in ("enabled", "threshold", "keep_recent", "summary_max_tokens"):
            if key in arguments:
                merged[key] = arguments[key]
                changed.append(f"{key} = {arguments[key]}")
        await self.set_conversation_state(_COMPRESSION_CONFIG_KEY, merged, conv_id)
        if changed:
            return "Updated compression settings:\n" + "\n".join(
                f"  {c}" for c in changed
            )
        return "No changes — no settings were provided."

    async def _tool_clear_compression(self, arguments: dict[str, Any]) -> str:
        conv_id = self._resolve_conversation_id(arguments.get("_conversation_id"))
        await self.clear_conversation_state(_COMPRESSION_STATE_KEY, conv_id)
        return (
            "Compression state cleared. The full message history is still "
            "stored — a new summary will be generated when the conversation "
            "exceeds the threshold again."
        )

    async def _tool_force_compression(self, arguments: dict[str, Any]) -> str:
        conv_id = self._resolve_conversation_id(arguments.get("_conversation_id"))
        if not self._backends:
            return "Cannot compress — no AI backend is available."
        messages = await self._load_conversation(conv_id)
        if len(messages) < 4:
            return "Not enough messages to compress (need at least 4)."

        await self._maybe_compress_history(messages, conv_id, force=True)

        state = await self.get_conversation_state(_COMPRESSION_STATE_KEY, conv_id)
        if isinstance(state, dict) and state.get("summary"):
            compressed_up_to = state.get("compressed_up_to", 0)
            remaining = len(messages) - compressed_up_to
            return (
                f"Compressed {compressed_up_to} of {len(messages)} messages "
                f"into a {len(state['summary']):,}-character summary. "
                f"{remaining} recent messages kept verbatim."
            )
        return "Compression failed — check the logs for details."

    # --- Persona tool handlers ---

    async def _tool_get_persona(self) -> str:
        persona_text = self._persona.persona if self._persona else DEFAULT_PERSONA
        return _json.dumps({"persona": persona_text})

    async def _tool_update_persona(self, arguments: dict[str, Any]) -> str:
        if self._persona is None:
            return _json.dumps({"error": "Persona not initialized"})
        text = arguments["text"]
        await self._persona.update_persona(text)
        return _json.dumps({"status": "updated", "length": len(text)})

    async def _tool_reset_persona(self) -> str:
        if self._persona is None:
            return _json.dumps({"error": "Persona not initialized"})
        await self._persona.reset_persona()
        return _json.dumps({"status": "reset"})

    # --- Memory tool handler ---

    async def _tool_memory_action(self, arguments: dict[str, Any]) -> str:
        if self._memory is None:
            return "Memory system not initialized."
        action = arguments.get("action", "")
        # Caller identity is injected into ``arguments`` by the tool executor
        # (both the AI-driven path in ``_execute_tool_calls`` and the slash
        # command path in ``_invoke_slash_command``). Fall back to the
        # contextvar for any callers that invoke this handler directly.
        user_id = arguments.get("_user_id") or get_current_user().user_id
        if user_id in ("system", "guest"):
            return "Memory requires an authenticated user."
        match action:
            case "remember":
                return await self._memory.remember(user_id, arguments)
            case "recall":
                return await self._memory.recall(user_id, arguments)
            case "update":
                return await self._memory.update(user_id, arguments)
            case "forget":
                return await self._memory.forget(user_id, arguments)
            case "list":
                return await self._memory.list_memories(user_id)
            case _:
                return f"Unknown memory action: {action}"

    async def _tool_rename_conversation(self, arguments: dict[str, Any]) -> str:
        title = arguments.get("title", "").strip()
        if not title:
            return _json.dumps({"error": "Title is required"})
        conv_id = get_current_conversation_id()
        if not conv_id or not self._storage:
            return _json.dumps({"error": "No active conversation"})

        data = await self._storage.get("ai_conversations", conv_id)
        if data is None:
            return _json.dumps({"error": "Conversation not found"})

        data["title"] = title
        await self._storage.put("ai_conversations", conv_id, data)

        # Emit event so WebSocket clients can update their UI
        if self._resolver:
            event_bus_svc = self._resolver.get_capability("event_bus")
            if event_bus_svc is not None:
                from gilbert.interfaces.events import Event, EventBusProvider

                if isinstance(event_bus_svc, EventBusProvider):
                    await event_bus_svc.bus.publish(
                        Event(
                            event_type="chat.conversation.renamed",
                            data={
                                "conversation_id": conv_id,
                                "title": title,
                            },
                            source="ai",
                        )
                    )

        return _json.dumps({"status": "renamed", "title": title})

    # --- Logging ---

    async def _record_round_usage(
        self,
        *,
        response: AIResponse,
        user_ctx: UserContext | None,
        conversation_id: str,
        profile: AIContextProfile | None,
        backend_name: str,
        round_num: int,
        turn_totals: dict[str, Any],
        invocation_source: str,
    ) -> dict[str, Any]:
        """Persist this round's usage (if a UsageRecorder is registered)
        and fold its numbers into ``turn_totals``.

        Returns a per-round usage dict (input/output/cache/cost) so the
        caller can attach it to a ``turn_rounds`` entry. Safe to call with
        a missing recorder or missing usage — returns a zeroed dict and
        leaves ``turn_totals`` untouched. Failures inside the recorder are
        logged and swallowed so the AI loop never breaks on reporting.
        """
        usage = response.usage
        if usage is None:
            return _empty_round_usage()

        # Fold raw token counts into the turn totals unconditionally. Cost
        # gets added after the recorder computes it (or, without a
        # recorder, remains 0).
        turn_totals["input_tokens"] += usage.input_tokens
        turn_totals["output_tokens"] += usage.output_tokens
        turn_totals["cache_creation_tokens"] += usage.cache_creation_tokens
        turn_totals["cache_read_tokens"] += usage.cache_read_tokens
        turn_totals["rounds"] += 1

        recorder = self._resolve_usage_recorder()
        cost = 0.0
        if recorder is not None:
            tool_names = [tc.tool_name for tc in response.message.tool_calls]
            try:
                rec: UsageRecord = await recorder.record_round(
                    user_ctx=user_ctx or UserContext.SYSTEM,
                    conversation_id=conversation_id,
                    profile=profile.name if profile is not None else "",
                    backend=backend_name,
                    model=response.model,
                    usage=usage,
                    tool_names=tool_names,
                    stop_reason=response.stop_reason.value,
                    round_num=round_num,
                    invocation_source=invocation_source,
                )
                cost = rec.cost_usd
            except Exception as exc:
                logger.warning(
                    "Usage recorder raised (conv=%s round=%d): %s",
                    conversation_id,
                    round_num,
                    exc,
                )

        turn_totals["cost_usd"] = round(turn_totals["cost_usd"] + cost, 6)

        return {
            "input_tokens": usage.input_tokens,
            "output_tokens": usage.output_tokens,
            "cache_creation_tokens": usage.cache_creation_tokens,
            "cache_read_tokens": usage.cache_read_tokens,
            "cost_usd": round(cost, 6),
        }

    def _resolve_usage_recorder(self) -> UsageRecorder | None:
        if self._resolver is None:
            return None
        svc = self._resolver.get_capability("usage_recording")
        if svc is None:
            return None
        if isinstance(svc, UsageRecorder):
            return svc
        return None

    def _log_api_call(self, request: AIRequest, response: AIResponse, round_num: int) -> None:
        usage_str = ""
        if response.usage:
            usage_str = f" tokens={response.usage.input_tokens}+{response.usage.output_tokens}"
        ai_logger.debug(
            "AI call round=%d model=%s stop=%s%s tools=%d messages=%d",
            round_num,
            response.model,
            response.stop_reason.value,
            usage_str,
            len(request.tools),
            len(request.messages),
        )

    # --- WebSocket RPC handlers ---

    @staticmethod
    def _filter_blocks_for_user(
        blocks: list[dict[str, Any]],
        user_id: str,
    ) -> list[dict[str, Any]]:
        """Filter UI blocks by for_user/exclude_user targeting."""
        return [
            b
            for b in blocks
            if (not b.get("for_user") or b.get("for_user") == user_id)
            and b.get("exclude_user") != user_id
        ]

    def get_ws_handlers(self) -> dict[str, Any]:
        return {
            "chat.message.send": self._ws_chat_send,
            "chat.message.cancel": self._ws_chat_cancel,
            "chat.form.submit": self._ws_form_submit,
            "chat.history.load": self._ws_history_load,
            "chat.conversation.list": self._ws_conversation_list,
            "chat.conversation.create": self._ws_conversation_create,
            "chat.conversation.rename": self._ws_conversation_rename,
            "chat.conversation.delete": self._ws_conversation_delete,
            "chat.room.create": self._ws_room_create,
            "chat.room.join": self._ws_room_join,
            "chat.room.leave": self._ws_room_leave,
            "chat.room.kick": self._ws_room_kick,
            "chat.room.invite": self._ws_room_invite,
            "chat.room.invite_revoke": self._ws_room_invite_revoke,
            "chat.room.invite_respond": self._ws_room_invite_respond,
            "chat.user.list": self._ws_chat_list_users,
            "slash.commands.list": self._ws_slash_commands_list,
            "chat.models.list": self._ws_models_list,
        }

    async def _ws_slash_commands_list(
        self,
        conn: Any,
        frame: dict[str, Any],
    ) -> dict[str, Any] | None:
        """Return the slash commands the caller can invoke.

        Drives chat-input autocomplete. Results are filtered by RBAC so
        users only see commands they're actually allowed to run.
        """

        slash_cmds = self._slash_commands_for_user(conn.user_ctx)
        commands: list[dict[str, Any]] = []
        for cmd_name, (provider, tool_def) in sorted(slash_cmds.items()):
            # ``cmd_name`` may contain a space (e.g. "radio start") or a
            # plugin-namespace dot (e.g. "currev.time_logs"); either way
            # it IS the full invocation so the usage string reflects the
            # grouped / namespaced form the user actually types.
            commands.append(
                {
                    "command": cmd_name,
                    "group": tool_def.slash_group or "",
                    "tool_name": tool_def.name,
                    "provider": provider.tool_provider_name,
                    "description": tool_def.description,
                    "help": tool_def.slash_help or tool_def.description,
                    "usage": format_usage(tool_def, full_command=cmd_name),
                    "required_role": tool_def.required_role,
                    "parameters": [
                        {
                            "name": p.name,
                            "type": p.type.value,
                            "description": p.description,
                            "required": p.required,
                            "default": p.default,
                            "enum": p.enum,
                        }
                        for p in tool_def.parameters
                        if not p.name.startswith("_")
                    ],
                }
            )
        return {
            "type": "slash.commands.list.result",
            "ref": frame.get("id"),
            "commands": commands,
        }

    async def _ws_models_list(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        """Return per-backend model lists for the chat UI."""
        return {
            "type": "chat.models.list.result",
            "ref": frame.get("id"),
            "backends": self.get_backends_with_models(),
        }

    async def _ws_chat_send(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:

        message = frame.get("message", "").strip()
        raw_attachments = frame.get("attachments") or []
        frame_model = str(frame.get("model") or "")
        frame_backend = str(frame.get("backend") or "")
        try:
            attachments = _parse_frame_attachments(raw_attachments)
        except ValueError as exc:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": str(exc),
                "code": 400,
            }
        if not message and not attachments:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "message is required",
                "code": 400,
            }

        # Register this RPC task in the in-flight registry keyed by the
        # WS frame id so ``chat.message.cancel`` can look it up and
        # interrupt the running turn. The registry entry also records
        # the originating user_id so cancel can enforce "only the
        # originator can interrupt." Removed in the finally block
        # below whether the turn finishes, errors, or is cancelled.
        rpc_ref = str(frame.get("id") or "")
        cancel_key = rpc_ref
        current_task = asyncio.current_task()
        if cancel_key and current_task is not None:
            self._in_flight_chats[cancel_key] = (
                current_task,
                conn.user_ctx.user_id,
            )

        conversation_id = frame.get("conversation_id") or None

        # Check if this is a shared room
        is_shared = False
        conv_data = None
        if conversation_id and self._storage:
            conv_data = await self._storage.get(_COLLECTION, conversation_id)
            if conv_data:
                is_shared = conv_data.get("shared", False)

        # Slash commands bypass the AI entirely and are handled inside
        # chat(). In shared rooms they also bypass the mentions_gilbert
        # check — invoking a tool is always intentional. We use the
        # longest-prefix matcher here so grouped commands like
        # ``/radio start`` are detected correctly.
        is_slash_command = False
        if extract_command_name(message) is not None:
            slash_cmds = self._slash_commands_for_user(conn.user_ctx)
            is_slash_command = self._match_slash_command(message, slash_cmds) is not None

        interrupted = False
        reply_model = ""
        reply_turn_usage: dict[str, Any] | None = None
        try:
            if is_shared:
                from gilbert.core.chat import build_room_context, mentions_gilbert, publish_event

                # ``is_shared`` was set from ``conv_data.get("shared")``
                # so we know conv_data is a dict at this point.
                assert conv_data is not None
                assert conversation_id is not None

                addressed = mentions_gilbert(message) or is_slash_command
                tagged_message = f"[{conn.user_ctx.display_name}]: {message}"

                response_text = ""
                ui_blocks: list[dict[str, Any]] = []
                tool_usage: list[dict[str, Any]] = []
                reply_attachments: list[FileAttachment] = []

                reply_rounds: list[dict[str, Any]] = []
                if addressed:
                    # Slash commands need the raw "/cmd ..." text so the
                    # parser recognizes them; the AI-chat path uses the
                    # tagged form so Gilbert knows who said what.
                    chat_message = message if is_slash_command else tagged_message
                    turn_result = await self.chat(
                        user_message=chat_message,
                        conversation_id=conversation_id,
                        user_ctx=conn.user_ctx,
                        system_prompt=build_room_context(conv_data, conn.user_ctx),
                        ai_profile=self._chat_profile,
                        attachments=attachments,
                        model=frame_model,
                        backend_override=frame_backend,
                    )
                    response_text = turn_result.response_text
                    conv_id = turn_result.conversation_id
                    ui_blocks = turn_result.ui_blocks
                    tool_usage = turn_result.tool_usage
                    reply_attachments = turn_result.attachments
                    reply_rounds = turn_result.rounds
                    interrupted = turn_result.interrupted
                    reply_model = turn_result.model
                    reply_turn_usage = turn_result.turn_usage
                else:
                    # Store message without invoking AI
                    conv_id = conversation_id
                    messages = await self._load_conversation(conversation_id)
                    messages.append(
                        Message(
                            role=MessageRole.USER,
                            content=tagged_message,
                            author_id=conn.user_ctx.user_id,
                            author_name=conn.user_ctx.display_name,
                            attachments=list(attachments),
                        )
                    )
                    await self._save_conversation(conv_id, messages, user_ctx=conn.user_ctx)

                # Broadcast to room members
                gilbert = conn.manager.gilbert
                if gilbert:
                    await publish_event(
                        gilbert,
                        "chat.message.created",
                        {
                            "conversation_id": conv_id,
                            "author_id": conn.user_ctx.user_id,
                            "author_name": conn.user_ctx.display_name,
                            "content": response_text,
                            "user_message": message,
                            "ui_blocks": ui_blocks,
                            "attachments": _serialize_attachments_for_wire(reply_attachments),
                        },
                    )
            else:
                # Personal chat — normal AI flow
                turn_result = await self.chat(
                    user_message=message,
                    conversation_id=conversation_id,
                    user_ctx=conn.user_ctx,
                    ai_profile=self._chat_profile,
                    attachments=attachments,
                    model=frame_model,
                    backend_override=frame_backend,
                )
                response_text = turn_result.response_text
                conv_id = turn_result.conversation_id
                ui_blocks = turn_result.ui_blocks
                tool_usage = turn_result.tool_usage
                reply_attachments = turn_result.attachments
                reply_rounds = turn_result.rounds
                interrupted = turn_result.interrupted
                reply_model = turn_result.model
                reply_turn_usage = turn_result.turn_usage
        except Exception as exc:
            logger.warning("chat.message.send failed", exc_info=True)
            return {"type": "gilbert.error", "ref": frame.get("id"), "error": str(exc), "code": 500}
        finally:
            # Always unregister, even on error. Cancel RPCs that arrive
            # after this won't find a matching task and will no-op.
            if cancel_key:
                self._in_flight_chats.pop(cancel_key, None)

        # Persist model preference when the user explicitly selects one,
        # and clear any stored preference when they explicitly send with
        # "Default" (empty selection). Without the clear path, once a
        # conversation gets a model preference pinned, picking "Default"
        # later is a no-op on the server — the pin survives on reload
        # and the chat keeps defaulting to whatever was previously
        # pinned (e.g. Haiku).
        if conv_id:
            if frame_model or frame_backend:
                await self.set_conversation_state(
                    "model_preference",
                    {"backend": frame_backend, "model": frame_model},
                    conv_id,
                )
            else:
                existing = await self.get_conversation_state(
                    "model_preference", conv_id
                )
                if existing:
                    await self.set_conversation_state(
                        "model_preference",
                        {"backend": "", "model": ""},
                        conv_id,
                    )

        return {
            "type": "chat.message.send.result",
            "ref": frame.get("id"),
            "response": response_text,
            "conversation_id": conv_id,
            "ui_blocks": self._filter_blocks_for_user(ui_blocks, conn.user_id),
            "tool_usage": tool_usage,
            "attachments": _serialize_attachments_for_wire(reply_attachments),
            "rounds": reply_rounds,
            "interrupted": interrupted,
            "model": reply_model,
            "turn_usage": reply_turn_usage,
        }

    async def _ws_chat_cancel(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        """Interrupt a running chat turn by its originating RPC ref.

        Frontend calls this when the user clicks the stop button while
        Gilbert is working. The frame carries the ``ref`` of the
        in-flight ``chat.message.send`` RPC; we look up the registered
        task in ``self._in_flight_chats`` and call ``.cancel()`` on it.
        ``AIService.chat()`` catches the resulting ``CancelledError``
        inside its agentic loop, marks the trailing assistant row as
        interrupted, shields the partial-state save from the
        cancellation, and returns a normal ``ChatTurnResult`` with
        ``interrupted=True``. The original ``chat.message.send`` RPC
        then resolves with an interrupted result frame — this cancel
        RPC just ACKs the request.

        Only the originator of the turn can cancel it. Anyone else who
        tries gets a 403. This matches the product decision: in a
        shared room, a user can't stop another user's in-flight turn.

        Cancels that arrive after the turn has already completed are
        treated as no-ops (``cancelled=False`` in the ack).
        """
        target_ref = str(frame.get("ref") or "")
        if not target_ref:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "ref is required",
                "code": 400,
            }

        entry = self._in_flight_chats.get(target_ref)
        if entry is None:
            return {
                "type": "chat.message.cancel.result",
                "ref": frame.get("id"),
                "cancelled": False,
                "reason": "not_found",
            }

        task, originator_id = entry
        if originator_id != conn.user_ctx.user_id:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Only the originator can cancel this turn.",
                "code": 403,
            }

        if task.done():
            return {
                "type": "chat.message.cancel.result",
                "ref": frame.get("id"),
                "cancelled": False,
                "reason": "already_done",
            }

        task.cancel()
        return {
            "type": "chat.message.cancel.result",
            "ref": frame.get("id"),
            "cancelled": True,
        }

    async def _ws_conversation_create(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        """Create an empty named personal conversation."""

        title = (frame.get("title") or "").strip() or "New conversation"

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        conv_id = str(uuid.uuid4())
        now = datetime.now(UTC).isoformat()

        await self._storage.put(
            _COLLECTION,
            conv_id,
            {
                "title": title,
                "user_id": conn.user_ctx.user_id,
                "messages": [],
                "created_at": now,
                "updated_at": now,
            },
        )

        return {
            "type": "chat.conversation.create.result",
            "ref": frame.get("id"),
            "conversation_id": conv_id,
            "title": title,
        }

    async def _ws_form_submit(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:

        conversation_id = frame.get("conversation_id")
        block_id = frame.get("block_id")
        values = frame.get("values", {})

        if not conversation_id or not block_id:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id and block_id required",
                "code": 400,
            }

        # Mark block as submitted in storage and check if shared room
        block_title = "Form"
        is_shared = False
        conv_data = None
        if self._storage is not None:
            conv_data = await self._storage.get(_COLLECTION, conversation_id)
            if conv_data:
                is_shared = conv_data.get("shared", False)
                for block in conv_data.get("ui_blocks", []):
                    if block.get("block_id") == block_id:
                        block["submitted"] = True
                        block["submission"] = values
                        block_title = block.get("title") or "Form"
                        break
                await self._storage.put(_COLLECTION, conversation_id, conv_data)

        # Build text message for AI
        form_message = f"[{conn.user_ctx.display_name} submitted: {block_title}]\n"
        for k, v in values.items():
            form_message += f"- {k}: {v}\n"

        try:
            system_prompt = None
            if is_shared and conv_data:
                from gilbert.core.chat import build_room_context

                system_prompt = build_room_context(conv_data, conn.user_ctx)

            turn_result = await self.chat(
                user_message=form_message,
                conversation_id=conversation_id,
                user_ctx=conn.user_ctx,
                system_prompt=system_prompt,
                ai_profile=self._chat_profile,
            )
            response_text = turn_result.response_text
            conv_id = turn_result.conversation_id
            ui_blocks = turn_result.ui_blocks
            reply_attachments = turn_result.attachments
            reply_rounds = turn_result.rounds
        except Exception as exc:
            logger.warning("chat.form.submit failed", exc_info=True)
            return {"type": "gilbert.error", "ref": frame.get("id"), "error": str(exc), "code": 500}

        # Broadcast to room members in shared rooms
        if is_shared:
            from gilbert.core.chat import publish_event

            gilbert = conn.manager.gilbert
            if gilbert:
                await publish_event(
                    gilbert,
                    "chat.message.created",
                    {
                        "conversation_id": conv_id,
                        "author_id": conn.user_ctx.user_id,
                        "author_name": conn.user_ctx.display_name,
                        "content": response_text,
                        "user_message": "",
                        "ui_blocks": ui_blocks,
                        "attachments": _serialize_attachments_for_wire(reply_attachments),
                    },
                )

        return {
            "type": "chat.form.submit.result",
            "ref": frame.get("id"),
            "response": response_text,
            "conversation_id": conv_id,
            "ui_blocks": self._filter_blocks_for_user(ui_blocks, conn.user_id),
            "attachments": _serialize_attachments_for_wire(reply_attachments),
            "rounds": reply_rounds,
        }

    async def _ws_history_load(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:

        conversation_id = frame.get("conversation_id")
        if not conversation_id:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id required",
                "code": 400,
            }

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        data = await self._storage.get(_COLLECTION, conversation_id)
        if data is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Conversation not found",
                "code": 404,
            }

        is_shared = data.get("shared", False)
        turns = self._group_persisted_messages_into_turns(
            data.get("messages", []),
            viewer_user_id=conn.user_id,
            include_author=is_shared,
        )

        ui_blocks = self._filter_blocks_for_user(
            data.get("ui_blocks", []),
            conn.user_id,
        )

        state = data.get("state", {})
        model_pref = state.get("model_preference") if isinstance(state, dict) else None

        result: dict[str, Any] = {
            "type": "chat.history.load.result",
            "ref": frame.get("id"),
            "turns": turns,
            "ui_blocks": ui_blocks,
            "shared": is_shared,
            "title": data.get("title", ""),
        }
        if model_pref:
            result["model_preference"] = model_pref
        if is_shared:
            result["members"] = data.get("members", [])
            result["invites"] = [
                {"user_id": inv["user_id"], "display_name": inv.get("display_name", "")}
                for inv in data.get("invites", [])
            ]
        return result

    def _group_persisted_messages_into_turns(
        self,
        rows: list[dict[str, Any]],
        viewer_user_id: str,
        include_author: bool,
    ) -> list[dict[str, Any]]:
        """Walk persisted message rows and emit one turn per user→assistant exchange.

        Each turn shape:
            {
              "user_message": {role, content, attachments, author_id, author_name},
              "rounds": [
                {"reasoning": str, "tools": [{tool_call_id, tool_name,
                                              arguments, result, is_error}]},
                ...
              ],
              "final_content": str,
              "final_attachments": [...],
              "final_author_id": str,        # only if include_author
              "final_author_name": str,      # only if include_author
              "incomplete": bool,            # true if turn never produced a
                                             # final assistant text (e.g. hit
                                             # max_tool_rounds)
            }

        The grouping logic mirrors what ``AIService.chat`` builds on the
        live path, so a refresh of an in-flight or completed conversation
        produces exactly the same turn structure that the live RPC
        result returned.

        ``viewer_user_id`` filters out messages targeted at other users via
        their ``visible_to`` list. ``include_author`` controls whether
        author fields are emitted on user messages and final assistant
        messages — only useful for shared rooms.
        """
        turns: list[dict[str, Any]] = []
        current: dict[str, Any] | None = None
        # Per-round building state. ``round_reasoning`` is the assistant
        # content from the row that emitted the most recent batch of
        # tool_calls; ``round_tools`` is the list of tool entries for
        # that round, indexed by tool_call_id so subsequent tool_result
        # rows can fill in the result/is_error fields. ``round_usage``
        # is that assistant row's ``usage`` dict (tokens + cost), carried
        # into the round entry so the chat UI can render per-round metrics.
        round_reasoning: str = ""
        round_tools: list[dict[str, Any]] = []
        round_tools_by_id: dict[str, dict[str, Any]] = {}
        round_usage: dict[str, Any] | None = None

        def finalize_round() -> None:
            """Push the in-progress round (if any) onto the current turn."""
            nonlocal round_reasoning, round_tools, round_tools_by_id, round_usage
            if current is None:
                return
            if not round_reasoning and not round_tools:
                return
            entry: dict[str, Any] = {
                "reasoning": round_reasoning,
                "tools": round_tools,
            }
            if round_usage:
                entry["usage"] = round_usage
            current["rounds"].append(entry)
            round_reasoning = ""
            round_tools = []
            round_tools_by_id = {}
            round_usage = None

        def start_turn(user_row: dict[str, Any]) -> None:
            """Open a new turn keyed on a user message row."""
            nonlocal current
            finalize_current_turn()
            current = {
                "user_message": self._build_turn_user_message(
                    user_row, include_author
                ),
                "rounds": [],
                "final_content": "",
                "final_attachments": [],
                "incomplete": False,
                "interrupted": False,
            }
            if include_author:
                current["final_author_id"] = ""
                current["final_author_name"] = ""

        def finalize_current_turn() -> None:
            """Close out the current turn and push it onto ``turns``."""
            nonlocal current
            if current is None:
                return
            finalize_round()
            # If we never saw a no-tool-calls assistant row, the turn
            # didn't reach a clean stopping point — usually because the
            # agentic loop hit max_tool_rounds or the AI errored. Mark it
            # so the frontend can render an "incomplete" indicator.
            # Slash commands that close the turn inline are always
            # considered complete regardless of final_content.
            slash_closed = bool(current.pop("_slash_closed", False))
            if (
                not slash_closed
                and not current["final_content"]
                and not current["final_attachments"]
            ):
                if current["rounds"]:
                    current["incomplete"] = True
            turns.append(current)
            current = None

        for row in rows:
            role = row.get("role")
            visible_to = row.get("visible_to")
            if visible_to is not None and viewer_user_id not in visible_to:
                continue

            if role == "user":
                start_turn(row)
                continue

            if role == "tool_result":
                if current is None:
                    # Tool result with no preceding user turn — orphan
                    # data, ignore.
                    continue
                for tr in row.get("tool_results", []) or []:
                    call_id = tr.get("tool_call_id", "")
                    entry = round_tools_by_id.get(call_id)
                    if entry is None:
                        # Result without a matching call — still surface
                        # it so the user can see something happened.
                        orphan = {
                            "tool_call_id": call_id,
                            "tool_name": "",
                            "is_error": bool(tr.get("is_error", False)),
                            "arguments": {},
                            "result": tr.get("content", ""),
                        }
                        round_tools.append(orphan)
                        round_tools_by_id[call_id] = orphan
                    else:
                        entry["result"] = tr.get("content", "")
                        entry["is_error"] = bool(tr.get("is_error", False))
                continue

            if role != "assistant":
                continue

            if current is None:
                # Assistant content with no preceding user turn — orphan
                # data, ignore.
                continue

            content = row.get("content", "") or ""
            tool_calls = row.get("tool_calls", []) or []
            inline_results = row.get("tool_results", []) or []

            if tool_calls:
                # Intermediate AI round (or slash-command row, which
                # carries tool_calls + tool_results inline). If we
                # already had a partially-built round, finalize it
                # before opening this one.
                if round_reasoning or round_tools:
                    finalize_round()

                # Strip the interrupt sentinel from round reasoning too
                # — if the user cancelled right between the assistant
                # row being appended and the tool executing, the row
                # carries both tool_calls and the marker. The marker
                # is for the AI, not the UI.
                row_interrupted = bool(row.get("interrupted"))
                round_reasoning = (
                    _strip_interrupt_marker(content) if row_interrupted else content
                )
                row_usage = row.get("usage")
                round_usage = row_usage if isinstance(row_usage, dict) else None
                if row_interrupted:
                    current["interrupted"] = True
                for tc in tool_calls:
                    call_id = tc.get("tool_call_id", "")
                    entry = {
                        "tool_call_id": call_id,
                        "tool_name": tc.get("tool_name", ""),
                        "is_error": False,
                        "arguments": self._sanitize_tool_args(
                            tc.get("arguments", {}) or {},
                        ),
                        "result": "",
                    }
                    round_tools.append(entry)
                    if call_id:
                        round_tools_by_id[call_id] = entry

                # Slash-command rows carry the tool_results on the same
                # row. Pair them up immediately so the round shows
                # complete data.
                for tr in inline_results:
                    call_id = tr.get("tool_call_id", "")
                    entry = round_tools_by_id.get(call_id)
                    if entry is not None:
                        entry["result"] = tr.get("content", "")
                        entry["is_error"] = bool(tr.get("is_error", False))

                # Slash-command rows ALSO carry the user-facing answer
                # text on the same row — the assistant ``content`` is
                # both the round narration AND the turn's final answer.
                # Treat such rows as turn-closing: emit the round and
                # set the final fields, then finalize the turn so the
                # next iteration starts fresh.
                if inline_results:
                    finalize_round()
                    current["final_content"] = content
                    current["final_attachments"] = self._serialize_persisted_attachments(
                        row.get("attachments")
                    )
                    # Slash-closing row — the turn is complete even if
                    # ``content`` is empty (e.g. the tool produced only
                    # UI blocks and we deliberately skipped persisting a
                    # JSON bubble). Without this flag the downstream
                    # incomplete-check would flag the turn incorrectly.
                    current["_slash_closed"] = True
                    if include_author:
                        current["final_author_id"] = row.get("author_id", "")
                        current["final_author_name"] = row.get("author_name", "")
                continue

            # Assistant row WITHOUT tool_calls — this is the final
            # answer for the current turn. Capture content + attachments
            # and finalize.
            finalize_round()
            # Strip the AI-facing interrupt sentinel from content before
            # setting final_content so the visible bubble only shows
            # whatever partial reply the user actually saw.
            display_content = (
                _strip_interrupt_marker(content) if row.get("interrupted") else content
            )
            current["final_content"] = display_content
            current["final_attachments"] = self._serialize_persisted_attachments(
                row.get("attachments")
            )
            # Capture the final round's usage (end_turn or max_tokens row
            # tokens) as ``final_usage``. The frontend reads both
            # ``rounds[].usage`` and ``final_usage`` to compute per-turn
            # totals on history replay.
            final_usage_raw = row.get("usage")
            if isinstance(final_usage_raw, dict):
                current["final_usage"] = final_usage_raw
            # Propagate the persisted interrupted marker: set on the
            # trailing assistant row by ``chat()`` when the user hit
            # stop mid-flight. Drives the subtle stop icon on the
            # frontend's TurnBubble.
            if row.get("interrupted"):
                current["interrupted"] = True
            if include_author:
                current["final_author_id"] = row.get("author_id", "")
                current["final_author_name"] = row.get("author_name", "")

        finalize_current_turn()
        # Sum per-round + final usage onto each turn so the chat UI can
        # render a single per-turn total without re-walking the shape.
        for turn in turns:
            turn["turn_usage"] = _sum_turn_usage(turn)
        return turns

    def _build_turn_user_message(
        self,
        row: dict[str, Any],
        include_author: bool,
    ) -> dict[str, Any]:
        """Project a persisted user row into the wire shape used by turns."""
        # Strip the [Name]: prefix from shared room content for display.
        # The prefix is stored for AI context but isn't user-visible.
        raw_content = row.get("content", "") or ""
        msg: dict[str, Any] = {
            "content": raw_content,
            "attachments": self._serialize_persisted_attachments(
                row.get("attachments"),
            ),
        }
        if include_author:
            msg["author_id"] = row.get("author_id", "")
            msg["author_name"] = row.get("author_name", "")
        # Surface legacy ``images`` field for old conversations as
        # inline image attachments — same behavior as the previous
        # message-list emit path.
        if not msg["attachments"]:
            legacy_images = row.get("images") or []
            if isinstance(legacy_images, list):
                for img in legacy_images:
                    if isinstance(img, dict) and img.get("data"):
                        msg["attachments"].append(
                            {
                                "kind": "image",
                                "name": "",
                                "media_type": img.get("media_type", ""),
                                "data": img.get("data"),
                            }
                        )
        return msg

    @staticmethod
    def _serialize_persisted_attachments(
        raw: Any,
    ) -> list[dict[str, Any]]:
        """Project the raw persisted attachments list onto the wire shape."""
        out: list[dict[str, Any]] = []
        if not isinstance(raw, list):
            return out
        for att in raw:
            if not isinstance(att, dict):
                continue
            kind = str(att.get("kind") or "")
            if not kind:
                continue
            entry: dict[str, Any] = {
                "kind": kind,
                "name": att.get("name", ""),
                "media_type": att.get("media_type", ""),
            }
            if att.get("data"):
                entry["data"] = att.get("data")
            if att.get("text"):
                entry["text"] = att.get("text")
            if att.get("workspace_skill"):
                entry["workspace_skill"] = att.get("workspace_skill")
            if att.get("workspace_path"):
                entry["workspace_path"] = att.get("workspace_path")
            if att.get("workspace_conv"):
                entry["workspace_conv"] = att.get("workspace_conv")
            out.append(entry)
        return out

    async def _ws_conversation_list(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        from gilbert.core.chat import conv_summary

        personal = await self.list_conversations(user_id=conn.user_id, limit=30)
        shared = await self.list_shared_conversations(user_id=conn.user_id, limit=30)

        conversations = [conv_summary(c, shared=True) for c in shared]
        conversations += [conv_summary(c, shared=False) for c in personal]

        return {
            "type": "chat.conversation.list.result",
            "ref": frame.get("id"),
            "conversations": conversations,
        }

    async def _ws_conversation_rename(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        from gilbert.core.chat import check_conversation_access, publish_event

        conversation_id = frame.get("conversation_id")
        title = (frame.get("title") or "").strip()
        if not conversation_id or not title:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id and title required",
                "code": 400,
            }

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        data = await self._storage.get(_COLLECTION, conversation_id)
        if data is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Conversation not found",
                "code": 404,
            }

        err = check_conversation_access(data, conn.user_ctx)
        if err:
            return {"type": "gilbert.error", "ref": frame.get("id"), "error": err, "code": 403}

        data["title"] = title
        await self._storage.put(_COLLECTION, conversation_id, data)

        gilbert = conn.manager.gilbert
        if gilbert is not None:
            await publish_event(
                gilbert,
                "chat.conversation.renamed",
                {"conversation_id": conversation_id, "title": title},
            )

        return {
            "type": "chat.conversation.rename.result",
            "ref": frame.get("id"),
            "status": "ok",
            "title": title,
        }

    async def _ws_conversation_delete(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:

        conversation_id = frame.get("conversation_id")
        if not conversation_id:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id required",
                "code": 400,
            }

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        data = await self._storage.get(_COLLECTION, conversation_id)
        if data is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Conversation not found",
                "code": 404,
            }
        if data.get("shared"):
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Use room destroy for shared conversations",
                "code": 400,
            }
        conv_owner = data.get("user_id", "")
        if conv_owner and conn.user_id != "system" and conv_owner != conn.user_id:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Access denied",
                "code": 403,
            }

        await self._storage.delete(_COLLECTION, conversation_id)
        # Tell subscribers (SkillService for workspace cleanup, etc.)
        # that this conversation is gone. Same event name as the room
        # destroy path so subscribers only need one handler.
        await self._publish_event(
            "chat.conversation.destroyed",
            {
                "conversation_id": conversation_id,
                "owner_id": conv_owner,
            },
        )
        return {"type": "chat.conversation.delete.result", "ref": frame.get("id"), "status": "ok"}

    async def _ws_room_create(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        from gilbert.core.chat import publish_event

        title = (frame.get("title") or "").strip()
        if not title:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "title required",
                "code": 400,
            }
        visibility = frame.get("visibility", "public")

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        import uuid as _uuid
        from datetime import datetime

        conv_id = str(_uuid.uuid4())
        now = datetime.now(UTC).isoformat()

        members = [
            {
                "user_id": conn.user_id,
                "display_name": conn.user_ctx.display_name,
                "role": "owner",
                "joined_at": now,
            }
        ]
        data = {
            "shared": True,
            "visibility": visibility,
            "title": title,
            "user_id": conn.user_id,
            "members": members,
            "messages": [],
            "created_at": now,
            "updated_at": now,
        }
        await self._storage.put(_COLLECTION, conv_id, data)

        gilbert = conn.manager.gilbert
        if gilbert is not None:
            await publish_event(
                gilbert,
                "chat.conversation.created",
                {
                    "conversation_id": conv_id,
                    "title": title,
                    "shared": True,
                    "members": members,
                    "visibility": visibility,
                },
            )

        return {
            "type": "chat.room.create.result",
            "ref": frame.get("id"),
            "conversation_id": conv_id,
            "title": title,
            "members": [
                {"user_id": m["user_id"], "display_name": m["display_name"]} for m in members
            ],
        }

    async def _ws_room_join(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        from gilbert.core.chat import publish_event

        conversation_id = frame.get("conversation_id")
        if not conversation_id:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id required",
                "code": 400,
            }

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        data = await self._storage.get(_COLLECTION, conversation_id)
        if data is None or not data.get("shared"):
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Room not found",
                "code": 404,
            }

        members = data.get("members", [])
        if any(m.get("user_id") == conn.user_id for m in members):
            return {
                "type": "chat.room.join.result",
                "ref": frame.get("id"),
                "status": "already_member",
            }

        from datetime import datetime

        members.append(
            {
                "user_id": conn.user_id,
                "display_name": conn.user_ctx.display_name,
                "role": "member",
                "joined_at": datetime.now(UTC).isoformat(),
            }
        )
        data["members"] = members
        await self._storage.put(_COLLECTION, conversation_id, data)

        gilbert = conn.manager.gilbert
        if gilbert is not None:
            await publish_event(
                gilbert,
                "chat.member.joined",
                {
                    "conversation_id": conversation_id,
                    "user_id": conn.user_id,
                    "display_name": conn.user_ctx.display_name,
                },
            )

        return {"type": "chat.room.join.result", "ref": frame.get("id"), "status": "ok"}

    async def _ws_room_leave(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        from gilbert.core.chat import publish_event

        conversation_id = frame.get("conversation_id")
        if not conversation_id:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id required",
                "code": 400,
            }

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        data = await self._storage.get(_COLLECTION, conversation_id)
        if data is None or not data.get("shared"):
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Room not found",
                "code": 404,
            }

        gilbert = conn.manager.gilbert

        # Owner leaving destroys the room
        if data.get("user_id") == conn.user_id:
            await self._storage.delete(_COLLECTION, conversation_id)
            if gilbert is not None:
                await publish_event(
                    gilbert, "chat.conversation.destroyed", {"conversation_id": conversation_id}
                )
            return {"type": "chat.room.leave.result", "ref": frame.get("id"), "status": "destroyed"}

        members = [m for m in data.get("members", []) if m.get("user_id") != conn.user_id]
        data["members"] = members
        await self._storage.put(_COLLECTION, conversation_id, data)
        if gilbert is not None:
            await publish_event(
                gilbert,
                "chat.member.left",
                {"conversation_id": conversation_id, "user_id": conn.user_id},
            )

        return {"type": "chat.room.leave.result", "ref": frame.get("id"), "status": "ok"}

    async def _ws_room_kick(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        from gilbert.core.chat import publish_event

        conversation_id = frame.get("conversation_id")
        target_user = frame.get("user_id")
        if not conversation_id or not target_user:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id and user_id required",
                "code": 400,
            }

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        data = await self._storage.get(_COLLECTION, conversation_id)
        if data is None or not data.get("shared"):
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Room not found",
                "code": 404,
            }
        if data.get("user_id") != conn.user_id:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Only the room owner can kick members",
                "code": 403,
            }

        members = [m for m in data.get("members", []) if m.get("user_id") != target_user]
        data["members"] = members
        await self._storage.put(_COLLECTION, conversation_id, data)

        gilbert = conn.manager.gilbert
        if gilbert is not None:
            await publish_event(
                gilbert,
                "chat.member.kicked",
                {"conversation_id": conversation_id, "user_id": target_user},
            )

        return {"type": "chat.room.kick.result", "ref": frame.get("id"), "status": "ok"}

    async def _ws_room_invite(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        from gilbert.core.chat import publish_event

        conversation_id = frame.get("conversation_id")
        user_ids = frame.get("user_ids", [])
        # Support single user_id for backwards compat
        if not user_ids and frame.get("user_id"):
            user_ids = [
                {"user_id": frame["user_id"], "display_name": frame.get("display_name", "")}
            ]
        if not conversation_id or not user_ids:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id and user_ids required",
                "code": 400,
            }

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        data = await self._storage.get(_COLLECTION, conversation_id)
        if data is None or not data.get("shared"):
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Room not found",
                "code": 404,
            }

        members = data.get("members", [])
        invites = data.get("invites", [])
        member_ids = {m.get("user_id") for m in members}
        invite_ids = {inv.get("user_id") for inv in invites}

        from datetime import datetime

        now = datetime.now(UTC).isoformat()
        invited = []

        for entry in user_ids:
            target_user = entry.get("user_id") if isinstance(entry, dict) else entry
            display_name = entry.get("display_name", "") if isinstance(entry, dict) else ""
            if target_user in member_ids or target_user in invite_ids:
                continue
            invites.append(
                {
                    "user_id": target_user,
                    "display_name": display_name,
                    "invited_by": conn.user_id,
                    "invited_at": now,
                }
            )
            invite_ids.add(target_user)
            invited.append({"user_id": target_user, "display_name": display_name})

        data["invites"] = invites
        await self._storage.put(_COLLECTION, conversation_id, data)

        gilbert = conn.manager.gilbert
        if gilbert is not None:
            for inv in invited:
                await publish_event(
                    gilbert,
                    "chat.invite.created",
                    {
                        "conversation_id": conversation_id,
                        "title": data.get("title", ""),
                        "user_id": inv["user_id"],
                        "display_name": inv["display_name"],
                        "invited_by": conn.user_id,
                        "invited_by_name": conn.user_ctx.display_name,
                    },
                )

        return {
            "type": "chat.room.invite.result",
            "ref": frame.get("id"),
            "status": "ok",
            "invited": invited,
        }

    async def _ws_room_invite_revoke(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:

        conversation_id = frame.get("conversation_id")
        target_user = frame.get("user_id")
        if not conversation_id or not target_user:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id and user_id required",
                "code": 400,
            }

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        data = await self._storage.get(_COLLECTION, conversation_id)
        if data is None or not data.get("shared"):
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Room not found",
                "code": 404,
            }

        invites = data.get("invites", [])
        data["invites"] = [inv for inv in invites if inv.get("user_id") != target_user]
        await self._storage.put(_COLLECTION, conversation_id, data)

        return {
            "type": "chat.room.invite_revoke.result",
            "ref": frame.get("id"),
            "status": "ok",
        }

    async def _ws_room_invite_respond(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        from gilbert.core.chat import publish_event

        conversation_id = frame.get("conversation_id")
        action = frame.get("action")  # "accept" or "decline"
        if not conversation_id or action not in ("accept", "decline"):
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "conversation_id and action (accept/decline) required",
                "code": 400,
            }

        if self._storage is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Storage not available",
                "code": 503,
            }

        data = await self._storage.get(_COLLECTION, conversation_id)
        if data is None or not data.get("shared"):
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Room not found",
                "code": 404,
            }

        invites = data.get("invites", [])
        invite = next((inv for inv in invites if inv.get("user_id") == conn.user_id), None)
        if invite is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "No pending invite found",
                "code": 404,
            }

        # Remove the invite
        data["invites"] = [inv for inv in invites if inv.get("user_id") != conn.user_id]

        gilbert = conn.manager.gilbert

        if action == "accept":
            from datetime import datetime

            members = data.get("members", [])
            members.append(
                {
                    "user_id": conn.user_id,
                    "display_name": conn.user_ctx.display_name,
                    "role": "member",
                    "joined_at": datetime.now(UTC).isoformat(),
                }
            )
            data["members"] = members
            await self._storage.put(_COLLECTION, conversation_id, data)

            if gilbert is not None:
                await publish_event(
                    gilbert,
                    "chat.member.joined",
                    {
                        "conversation_id": conversation_id,
                        "user_id": conn.user_id,
                        "display_name": conn.user_ctx.display_name,
                    },
                )
        else:
            await self._storage.put(_COLLECTION, conversation_id, data)

            if gilbert is not None:
                await publish_event(
                    gilbert,
                    "chat.invite.declined",
                    {
                        "conversation_id": conversation_id,
                        "user_id": conn.user_id,
                    },
                )

        return {
            "type": "chat.room.invite_respond.result",
            "ref": frame.get("id"),
            "status": "ok",
            "action": action,
        }

    async def _ws_chat_list_users(
        self, conn: WsConnectionBase, frame: dict[str, Any]
    ) -> dict[str, Any] | None:
        """List all users for invite modal."""

        gilbert = conn.manager.gilbert
        if gilbert is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "Service unavailable",
                "code": 503,
            }

        user_svc = gilbert.service_manager.get_by_capability("users")
        if user_svc is None:
            return {
                "type": "gilbert.error",
                "ref": frame.get("id"),
                "error": "User service unavailable",
                "code": 503,
            }

        users = await user_svc.list_users(limit=200)
        user_list = [
            {
                "user_id": u.get("_id", ""),
                "display_name": u.get("display_name", u.get("username", "")),
            }
            for u in users
            if u.get("_id") != "system"
        ]

        return {"type": "chat.user.list.result", "ref": frame.get("id"), "users": user_list}
